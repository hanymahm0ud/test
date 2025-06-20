import sys
import random
import string
import os
import io
import sqlite3
import json
import openpyxl
import ctypes
import pymysql


from openpyxl.styles import Font, PatternFill, Alignment
from PyQt5.QtWidgets import QMessageBox
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QTableWidget,
    QTableWidgetItem, QPushButton, QLabel, QFrame, QHeaderView, QDialog,
    QFormLayout, QLineEdit, QComboBox, QDateEdit, QMessageBox, QDialogButtonBox,
    QTextEdit, QGroupBox, QGridLayout, QListWidget, QCheckBox,QListWidgetItem,QFileDialog)

from PyQt5.QtPrintSupport import QPrintDialog, QPrinter
from PyQt5.QtCore import Qt, QDate,QRect
from PyQt5.QtGui import QFont, QColor, QDoubleValidator, QIntValidator, QPainter,QIcon

from datetime import datetime, timedelta

if os.name == 'nt':  # Windows
    myappid = 'mycompany.myproduct.subproduct.version'  # أي معرف فريد
    ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
    
if hasattr(sys, 'frozen'):
    # إذا كان البرنامج مجمع (exe)
    try:
        if sys.stdout:
            sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
        if sys.stderr:
            sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
    except (AttributeError, IOError):
        pass
else:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

class LoginDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_app = parent
        self.current_user = None
        self.setWindowTitle('تسجيل الدخول')
        self.setModal(True)
        self.resize(500, 400)
        self.setStyleSheet("""
            QDialog {
                background-color: #f7fafc;
            }
            QLabel {
                font-size: 14px;
                color: #2d3748;
            }
            QLineEdit {
                padding: 10px;
                border: 1px solid #cbd5e0;
                border-radius: 5px;
                font-size: 14px;
            }
            QPushButton {
                padding: 10px;
                border-radius: 5px;
                font-size: 14px;
                font-weight: bold;
            }
        """)
        self.setupUI()
        
    def setupUI(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(50, 50, 50, 50)
        layout.setSpacing(25)
        
        # خلفية متدرجة للنافذة
        self.setStyleSheet("""
            QDialog {
                background: qlineargradient(x1: 0, y1: 0, x2: 1, y2: 1,
                    stop: 0 #667eea, stop: 1 #764ba2);
            }
        """)
        
        # العنوان الرئيسي
        title_label = QLabel("إدارة شركة النقل الثقيل")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("""
            QLabel {
                font-size: 28px;
                font-weight: bold;
                color: white;
                margin-bottom: 30px;
                text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
            }
        """)
        layout.addWidget(title_label)
        
        # مجموعة النموذج مع خلفية شفافة
        form_group = QGroupBox()
        form_group.setStyleSheet("""
            QGroupBox {
                background: rgba(255, 255, 255, 0.95);
                border-radius: 15px;
                padding: 20px;
                box-shadow: 0 8px 32px rgba(0,0,0,0.1);
            }
        """)
        form_layout = QGridLayout(form_group)
        form_layout.setSpacing(20)
        
        # حقل اسم المستخدم
        username_label = QLabel("اسم المستخدم:")
        username_label.setStyleSheet("""
            QLabel {
                font-size: 14px;
                font-weight: 600;
                color: #2d3748;
                margin-bottom: 5px;
            }
        """)
        
        self.username = QLineEdit()
        self.username.setPlaceholderText("أدخل اسم المستخدم")
        self.username.setStyleSheet("""
            QLineEdit {
                padding: 12px 15px;
                border: 2px solid #e2e8f0;
                border-radius: 8px;
                font-size: 14px;
                background: white;
                color: #2d3748;
            }
            QLineEdit:focus {
                border-color: #667eea;
                outline: none;
            }
            QLineEdit:hover {
                border-color: #cbd5e0;
            }
        """)
        
        form_layout.addWidget(username_label, 0, 1)
        form_layout.addWidget(self.username, 0, 0)
        
        # حقل كلمة المرور
        password_label = QLabel("كلمة المرور:")
        password_label.setStyleSheet("""
            QLabel {
                font-size: 14px;
                font-weight: 600;
                color: #2d3748;
                margin-bottom: 5px;
            }
        """)
        
        self.password = QLineEdit()
        self.password.setEchoMode(QLineEdit.Password)
        self.password.setPlaceholderText("أدخل كلمة المرور")
        self.password.setStyleSheet("""
            QLineEdit {
                padding: 12px 15px;
                border: 2px solid #e2e8f0;
                border-radius: 8px;
                font-size: 14px;
                background: white;
                color: #2d3748;
            }
            QLineEdit:focus {
                border-color: #667eea;
                outline: none;
            }
            QLineEdit:hover {
                border-color: #cbd5e0;
            }
        """)
        
        form_layout.addWidget(password_label, 1, 1)
        form_layout.addWidget(self.password, 1, 0)
        
        layout.addWidget(form_group)
        
        # الأزرار مع تصميم عصري
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.validate_login)
        button_box.rejected.connect(self.reject)
        
        ok_button = button_box.button(QDialogButtonBox.Ok)
        ok_button.setText("تسجيل الدخول")
        cancel_button = button_box.button(QDialogButtonBox.Cancel)
        cancel_button.setText("إلغاء")
        
        button_box.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #667eea, stop: 1 #764ba2);
                color: white;
                border: none;
                border-radius: 8px;
                padding: 12px 24px;
                font-size: 14px;
                font-weight: 600;
                min-width: 120px;
            }
            QPushButton:hover {
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #5a67d8, stop: 1 #6b46c1);
                transform: translateY(-1px);
            }
            QPushButton:pressed {
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #4c51bf, stop: 1 #553c9a);
                transform: translateY(0px);
            }
            QPushButton[text="إلغاء"] {
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #f56565, stop: 1 #e53e3e);
            }
            QPushButton[text="إلغاء"]:hover {
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #e53e3e, stop: 1 #c53030);
            }
            QPushButton[text="إلغاء"]:pressed {
                background: qlineargradient(x1: 0, y1: 0, x2: 0, y2: 1,
                    stop: 0 #c53030, stop: 1 #9c2222);
            }
        """)
        
        layout.addWidget(button_box)
        layout.addStretch()

    def validate_login(self):
        username = self.username.text()
        password = self.password.text()
        conn = pymysql.connect(self.parent_app.db_file)
        cursor = conn.cursor()
        cursor.execute("SELECT username, password, type, permissions FROM users WHERE username = %s", (username,))
        user = cursor.fetchone()
        conn.close()
        if user and user[1] == password:
            self.current_user = {
                'username': user[0],
                'password': user[1],
                'type': user[2],
                'permissions': json.loads(user[3])
            }
            print(user[2])
            self.accept()
        else:
            QMessageBox.warning(self, 'خطأ', 'اسم المستخدم أو كلمة المرور غير صحيحة!')

class UserManagementDialog(QDialog):
    def __init__(self, users, parent=None):
        super().__init__(parent)
        self.users = users
        self.parent_app = parent
        self.db_file = {
             "host": "localhost",
            "user": "hany",
            "password": "hany",
            "database": "transport_db"
        }        
        self.setWindowTitle('إدارة المستخدمين')
        self.setModal(True)
        self.resize(600, 500)
        self.setupUI()
        
    def setupUI(self):
        layout = QVBoxLayout(self)
        self.users_list = QListWidget()
        self.update_users_list()
        layout.addWidget(QLabel("قائمة المستخدمين:"))
        layout.addWidget(self.users_list)
        
        buttons_layout = QHBoxLayout()
        add_btn = QPushButton("إضافة مستخدم")
        add_btn.clicked.connect(self.add_user)
        buttons_layout.addWidget(add_btn)
        
        edit_btn = QPushButton("تعديل مستخدم")
        edit_btn.clicked.connect(self.edit_user)
        buttons_layout.addWidget(edit_btn)
        
        delete_btn = QPushButton("حذف مستخدم")
        delete_btn.clicked.connect(self.delete_user)
        buttons_layout.addWidget(delete_btn)
        
        layout.addLayout(buttons_layout)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        
    def update_users_list(self):
        self.users_list.clear()
        for user in self.users:
            self.users_list.addItem(f"{user['username']} ({user['type']})")
            
    def save_users_data(self):
        """حفظ بيانات المستخدمين في قاعدة البيانات"""
        try:
            conn = pymysql.connect(**self.db_file)
            cursor = conn.cursor()
            
            cursor.execute("DELETE FROM users")
            
            for user in self.users:
                cursor.execute(
                    "INSERT INTO users (username, password, type, permissions) VALUES (%s, %s, %s, %s)",
                    (user['username'], user['password'], 
                    user['type'], json.dumps(user['permissions']))
                )
            
            conn.commit()
            conn.close()
            QMessageBox.information(self, 'نجح', 'تم حفظ بيانات المستخدمين بنجاح!')
            
        except sqlite3.Error as e:
            print(f"خطأ في قاعدة البيانات: {str(e)}")
            QMessageBox.critical(self, 'خطأ', f'حدث خطأ في قاعدة البيانات: {str(e)}')
            if 'conn' in locals():
                conn.rollback()
                conn.close()
        except Exception as e:
            print(f"خطأ غير متوقع: {str(e)}")
            QMessageBox.critical(self, 'خطأ', f'حدث خطأ غير متوقع: {str(e)}')

    def add_user(self):
        dialog = UserEditDialog(self.users, self)
        if dialog.exec_() == QDialog.Accepted:
            user_data = dialog.get_user_data()
            if user_data['username'] and user_data['password']:
                if not any(user['username'] == user_data['username'] for user in self.users):
                    self.users.append(user_data)
                    self.save_users_data()  # حفظ في قاعدة البيانات
                    self.update_users_list()
                    QMessageBox.information(self, 'نجح', 'تم إضافة المستخدم بنجاح!')

    def edit_user(self):
        current_item = self.users_list.currentItem()
        if current_item:
            username = current_item.text().split(' (')[0]
            user = next(u for u in self.users if u['username'] == username)
            dialog = UserEditDialog(self.users, self, user)
            if dialog.exec_() == QDialog.Accepted:
                user_data = dialog.get_user_data()
                if user_data['username'] and user_data['password']:
                    user.update(user_data)
                    self.save_users_data()  # حفظ في قاعدة البيانات
                    self.update_users_list()
                    QMessageBox.information(self, 'نجح', 'تم تعديل المستخدم بنجاح!')

    def delete_user(self):
        current_item = self.users_list.currentItem()
        if current_item:
            username = current_item.text().split(' (')[0]
            if username == 'admin':
                QMessageBox.warning(self, 'خطأ', 'لا يمكن حذف المستخدم الافتراضي (admin)!')
                return
            reply = QMessageBox.question(self, 'تأكيد الحذف', f'هل تريد حذف "{username}"؟')
            if reply == QMessageBox.Yes:
                self.users[:] = [u for u in self.users if u['username'] != username]
                self.save_users_data()  # حفظ في قاعدة البيانات
                self.update_users_list()
                QMessageBox.information(self, 'نجح', 'تم حذف المستخدم بنجاح!')

class UserEditDialog(QDialog):
    def __init__(self, users, parent=None, user=None):
        super().__init__(parent)
        self.users = users
        self.user = user
        self.setWindowTitle('إضافة/تعديل مستخدم')
        self.setModal(True)
        self.resize(400, 400)
        self.setupUI()
        
    def setupUI(self):
        layout = QVBoxLayout(self)
        form_group = QGroupBox("بيانات المستخدم")
        form_layout = QFormLayout(form_group)
        
        self.username = QLineEdit()
        if self.user:
            self.username.setText(self.user['username'])
        form_layout.addRow("اسم المستخدم:", self.username)
        
        self.password = QLineEdit()
        if self.user:
            self.password.setText(self.user['password'])
        self.password.setEchoMode(QLineEdit.Password)
        form_layout.addRow("كلمة المرور:", self.password)
        
        self.user_type = QComboBox()
        self.user_type.addItems(["Admin", "User"])
        if self.user:
            self.user_type.setCurrentText(self.user['type'])
        form_layout.addRow("نوع المستخدم:", self.user_type)
        
        layout.addWidget(form_group)
        
        permissions_group = QGroupBox("الصلاحيات")
        permissions_layout = QVBoxLayout(permissions_group)
        
        self.permissions = {
        'trips': QCheckBox("الشحنات"),
        'drivers': QCheckBox("السائقين"),
        'countries': QCheckBox("إدارة الدول"),
        'trucks': QCheckBox("الشاحنات"),
        'companies': QCheckBox("الشركات"),
        'expenses': QCheckBox("المصاريف"),
        'reports': QCheckBox("التقارير"),
        'gallery': QCheckBox("معرض الشحنات")  # إضافة الصلاحية الجديدة
    }
        
        if self.user:
            for key, checkbox in self.permissions.items():
                checkbox.setChecked(self.user['permissions'].get(key, False))
                
        for checkbox in self.permissions.values():
            permissions_layout.addWidget(checkbox)
            
        layout.addWidget(permissions_group)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        
    def get_user_data(self):
        return {
            'username': self.username.text(),
            'password': self.password.text(),
            'type': self.user_type.currentText(),
            'permissions': {key: checkbox.isChecked() for key, checkbox in self.permissions.items()}
        }

class CountryManagerDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        
        self.db_file = {
             "host": "localhost",
            "user": "hany",
            "password": "hany",
            "database": "transport_db"
        } 
        self.countries_data = parent.load_countries_data()
        self.parent_app = parent
        self.drivers_data = parent.drivers_data
        self.setWindowTitle('إدارة الدول')
        self.setModal(True)
        self.resize(600, 500)
        self.countries = parent.countries_data.copy()  # استخدام نسخة من البيانات للعمل عليها
        self.setupUI()

    def setupUI(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)

        # عنوان
        title = QLabel("إدارة الدول")
        title.setStyleSheet("font-size: 24px; font-weight: bold; color: #2d3748; margin: 20px 0;")
        layout.addWidget(title)

        # مربع البحث
        search_layout = QHBoxLayout()
        search_label = QLabel("البحث:")
        search_label.setStyleSheet("font-size: 14px; font-weight: bold; color: #2d3748;")
        search_layout.addWidget(search_label)
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("ابحث عن دولة...")
        self.search_input.textChanged.connect(self.search_countries)
        self.search_input.setStyleSheet("""
            QLineEdit {
                padding: 8px;
                border: 2px solid #cbd5e0;
                border-radius: 5px;
                font-size: 14px;
                background-color: white;
            }
            QLineEdit:focus {
                border-color: #4299e1;
                outline: none;
            }
        """)
        search_layout.addWidget(self.search_input)
        
        # زر مسح البحث
        clear_search_btn = QPushButton("مسح")
        clear_search_btn.clicked.connect(self.clear_search)
        clear_search_btn.setStyleSheet("""
            QPushButton {
                background-color: #718096;
                color: white;
                padding: 8px 15px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #4a5568;
            }
        """)
        search_layout.addWidget(clear_search_btn)
        
        search_layout.addStretch()
        layout.addLayout(search_layout)

        # أزرار الأدوات
        toolbar_layout = QHBoxLayout()
        toolbar_layout.addStretch()

        add_btn = QPushButton("إضافة دولة")
        add_btn.clicked.connect(self.add_country)
        add_btn.setStyleSheet("background-color: #4299e1; color: white; padding: 10px 20px; border: none; border-radius: 5px; font-weight: bold;")
        toolbar_layout.addWidget(add_btn)

        edit_btn = QPushButton("تعديل دولة")
        edit_btn.clicked.connect(self.edit_country)
        edit_btn.setStyleSheet("background-color: #ecc94b; color: white; padding: 10px 20px; border: none; border-radius: 5px; font-weight: bold;")
        toolbar_layout.addWidget(edit_btn)

        delete_btn = QPushButton("حذف دولة")
        delete_btn.clicked.connect(self.delete_country)
        delete_btn.setStyleSheet("background-color: #e53e3e; color: white; padding: 10px 20px; border: none; border-radius: 5px; font-weight: bold;")
        toolbar_layout.addWidget(delete_btn)

        print_btn = QPushButton("طباعة")
        print_btn.clicked.connect(lambda: self.parent_app.print_table(self.countries_table, "قائمة الدول"))
        print_btn.setStyleSheet("background-color: #38a169; color: white; padding: 10px 20px; border: none; border-radius: 5px; font-weight: bold;")
        toolbar_layout.addWidget(print_btn)

        export_btn = QPushButton("تصدير إلى Excel")
        export_btn.clicked.connect(lambda: self.parent_app.export_table_to_excel_openpyxl(self.countries_table, "بيانات_الدول.xlsx"))
        export_btn.setStyleSheet("background-color: #ecc94b; color: white; padding: 10px 20px; border: none; border-radius: 5px; font-weight: bold;")
        toolbar_layout.addWidget(export_btn)

        layout.addLayout(toolbar_layout)

        # إنشاء الجدول
        self.countries_table = QTableWidget()
        self.countries_table.setColumnCount(2)
        self.countries_table.setHorizontalHeaderLabels(["اسم الدولة", "السائقون المسموح لهم"])
        self.countries_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.countries_table.horizontalHeader().setStretchLastSection(True)
        self.countries_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.countries_table.setAlternatingRowColors(True)
        self.countries_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.update_countries_table()
        layout.addWidget(self.countries_table)

        # زر إغلاق
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def search_countries(self):
        """البحث في الدول"""
        search_text = self.search_input.text().lower()
        
        for row in range(self.countries_table.rowCount()):
            # البحث في اسم الدولة
            country_name = self.countries_table.item(row, 0)
            if country_name:
                country_text = country_name.text().lower()
                # إظهار أو إخفاء الصف حسب نتيجة البحث
                if search_text in country_text:
                    self.countries_table.setRowHidden(row, False)
                else:
                    self.countries_table.setRowHidden(row, True)
            else:
                self.countries_table.setRowHidden(row, True)

    def clear_search(self):
        """مسح البحث وإظهار جميع الدول"""
        self.search_input.clear()
        for row in range(self.countries_table.rowCount()):
            self.countries_table.setRowHidden(row, False)

    def update_countries_table(self):
        self.countries_table.setRowCount(len(self.countries))
        for row, country in enumerate(self.countries):
            self.countries_table.setItem(row, 0, QTableWidgetItem(country['name']))
            drivers_text = ", ".join(country['allowed_drivers']) if country['allowed_drivers'] else "لا يوجد"
            self.countries_table.setItem(row, 1, QTableWidgetItem(drivers_text))

    
    def add_country(self):
        dialog = AddCountryDialog(self.drivers_data, self)
        if dialog.exec_() == QDialog.Accepted:
            country_data = dialog.get_country_data()
            if country_data['name'].strip():
                # التحقق من عدم وجود دولة بنفس الاسم
                if any(c['name'].lower() == country_data['name'].lower() for c in self.countries):
                    QMessageBox.warning(self, 'تحذير', 'هذه الدولة موجودة بالفعل!')
                    return
                    
                self.countries.append(country_data)
                self.update_countries_table()
                self.save_countries_data()  # Save to database immediately
                QMessageBox.information(self, 'نجح', 'تم إضافة الدولة بنجاح!')

    def edit_country(self):
        selected_row = self.countries_table.currentRow()
        if selected_row >= 0:
            country_data = self.countries[selected_row]
            dialog = AddCountryDialog(self.drivers_data, self)
            dialog.name.setText(country_data['name'])
            
            # تحديد السائقين المسموح لهم
            for i in range(dialog.allowed_drivers.count()):
                item = dialog.allowed_drivers.item(i)
                driver_name = item.text()
                item.setSelected(driver_name in country_data['allowed_drivers'])
                
            if dialog.exec_() == QDialog.Accepted:
                new_data = dialog.get_country_data()
                if new_data['name'].strip():
                    # التحقق من عدم وجود دولة أخرى بنفس الاسم (إذا تغير الاسم)
                    if new_data['name'] != country_data['name'] and \
                    any(c['name'].lower() == new_data['name'].lower() for c in self.countries):
                        QMessageBox.warning(self, 'تحذير', 'هذا الاسم موجود بالفعل!')
                        return
                        
                    self.countries[selected_row] = new_data
                    self.update_countries_table()
                    self.save_countries_data()
                    QMessageBox.information(self, 'نجح', 'تم تعديل الدولة بنجاح!')

    def delete_country(self):
        selected_row = self.countries_table.currentRow()
        if selected_row >= 0:
            country_name = self.countries[selected_row]['name']
            
            # التحقق من عدم وجود شحنات مرتبطة بهذه الدولة
            related_trips = [
                trip for trip in self.parent_app.trips_data 
                if trip['final_destination'] == country_name or 
                trip['intermediate_country'] == country_name or
                trip['start'] == country_name
            ]
            
            if related_trips:
                QMessageBox.warning(
                    self, 
                    'خطأ', 
                    f'لا يمكن حذف {country_name} لأنها مرتبطة ببعض الشحنات!'
                )
                return
                
            reply = QMessageBox.question(
                self, 
                'تأكيد الحذف', 
                f"هل تريد حذف {country_name}؟"
            )
            if reply == QMessageBox.Yes:
                del self.countries[selected_row]
                self.update_countries_table()
                self.save_countries_data()  # Save to database immediately
                QMessageBox.information(self, 'نجح', 'تم حذف الدولة بنجاح!')
    def save_countries_data(self):
        """حفظ بيانات الدول في قاعدة البيانات"""
        try:
            print("جاري اتصال بقاعدة البيانات...")
            conn = pymysql.connect(**self.db_file)
            cursor = conn.cursor()
            
            
            print("جاري إنشاء/تحديث جدول الدول...")
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS countries (
                    name TEXT PRIMARY KEY,
                    allowed_drivers TEXT
                )
            """)
            
            print("جاري حذف البيانات القديمة...")
            cursor.execute("DELETE FROM countries")
            
            print("جاري إدراج البيانات الجديدة...")
            for country in self.countries:  # تغيير من self.countries_data إلى self.countries
                cursor.execute(
                    "INSERT INTO countries (name, allowed_drivers) VALUES (%s, %s)",
                    (country['name'], json.dumps(country['allowed_drivers']))
                )
            
            conn.commit()
            conn.close()
            print("تم حفظ بيانات الدول بنجاح في قاعدة البيانات")
            
        except sqlite3.Error as e:
            print(f"خطأ في sqlite: {str(e)}")
            QMessageBox.critical(self, 'خطأ', f'حدث خطأ في قاعدة البيانات: {str(e)}')
            if 'conn' in locals():
                conn.rollback()
                conn.close()
                
        except Exception as e:
            print(f"خطأ غير متوقع: {str(e)}")
            QMessageBox.critical(self, 'خطأ', f'حدث خطأ غير متوقع: {str(e)}')

    def accept(self):
        """حفظ التغييرات عند الضغط على موافق"""
        try:
            # التحقق من عدم وجود أسماء دول مكررة
            country_names = [c['name'] for c in self.countries]
            if len(country_names) != len(set(country_names)):
                QMessageBox.warning(self, 'خطأ', 'يوجد دول بأسماء مكررة!')
                return
                
            print("جاري حفظ بيانات الدول...")
            
            # حفظ في قاعدة البيانات
            self.save_countries_data()
            
            # حفظ البيانات في التطبيق الرئيسي
            self.parent_app.countries_data = self.countries.copy()
            self.parent_app.countries = [c['name'] for c in self.countries]
            
            print("تم حفظ بيانات الدول بنجاح")
            super().accept()
            
        except Exception as e:
            print(f"حدث خطأ أثناء حفظ الدول: {str(e)}")
            QMessageBox.critical(self, 'خطأ', f'حدث خطأ أثناء الحفظ: {str(e)}')


class AddCountryDialog(QDialog):
    def __init__(self, drivers_data, parent=None):
        super().__init__(parent)
        self.drivers_data = drivers_data
        self.setWindowTitle('إضافة/تعديل دولة')
        self.setModal(True)
        self.resize(400, 400)
        self.setupUI()

    def setupUI(self):
        layout = QVBoxLayout(self)
        form_group = QGroupBox("بيانات الدولة")
        form_layout = QGridLayout(form_group)

        self.name = QLineEdit()
        form_layout.addWidget(QLabel("اسم الدولة:"), 0, 1)
        form_layout.addWidget(self.name, 0, 0)

        self.allowed_drivers = QListWidget()
        self.allowed_drivers.setSelectionMode(QListWidget.MultiSelection)
        
        # تحميل أسماء السائقين مع التحقق من وجودهم
        if self.drivers_data:
            for driver in self.drivers_data:
                if 'name' in driver:
                    item = QListWidgetItem(driver['name'])
                    self.allowed_drivers.addItem(item)
        
        form_layout.addWidget(QLabel("السائقون المسموح لهم:"), 1, 1)
        form_layout.addWidget(self.allowed_drivers, 1, 0)

        layout.addWidget(form_group)

        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.validate_and_accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)

    def validate_and_accept(self):
        """التحقق من صحة البيانات قبل القبول"""
        if not self.name.text().strip():
            QMessageBox.warning(self, 'خطأ', 'يرجى إدخال اسم الدولة!')
            return
            
        super().accept()

    def get_country_data(self):
        selected_drivers = []
        for i in range(self.allowed_drivers.count()):
            item = self.allowed_drivers.item(i)
            if item.isSelected():
                selected_drivers.append(item.text())
                
        return {
            'name': self.name.text().strip(),
            'allowed_drivers': selected_drivers
        }
    
class AddDriverDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('إضافة سائق جديد')
        self.setModal(True)
        self.resize(400, 300)
        self.setupUI()
        
    def setupUI(self):
        layout = QVBoxLayout(self)
        
        # إنشاء مجموعة بيانات السائق
        form_group = QGroupBox("بيانات السائق")
        grid_layout = QGridLayout(form_group)
        
        # إنشاء الحقول مع تحديد مواقعها في الGrid
        self.name = QLineEdit()
        name_label = QLabel("اسم السائق:*")  # إضافة علامة * للحقل الإلزامي
        name_label.setStyleSheet("color: #444; font-weight: bold;")  # تنسيق العنوان
        grid_layout.addWidget(name_label, 0, 1)
        grid_layout.addWidget(self.name, 0, 0)
        
        self.address = QLineEdit()
        grid_layout.addWidget(QLabel("العنوان:"), 1, 1)
        grid_layout.addWidget(self.address, 1, 0)
        
        self.age = QLineEdit()
        self.age.setValidator(QIntValidator(18, 70))
        grid_layout.addWidget(QLabel("السن:"), 2, 1)
        grid_layout.addWidget(self.age, 2, 0)
        
        self.national_id = QLineEdit()
        grid_layout.addWidget(QLabel("الرقم القومي:"), 3, 1)
        grid_layout.addWidget(self.national_id, 3, 0)
        
        # جعل عمود الإدخال أكثر اتساعاً
        grid_layout.setColumnStretch(0, 2)
        grid_layout.setColumnStretch(1, 1)
        
        layout.addWidget(form_group)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.validate_and_accept)  # تغيير الاتصال إلى دالة التحقق
        button_box.rejected.connect(self.reject)
        
        # تغيير نص الأزرار إلى العربية
        button_box.button(QDialogButtonBox.Ok).setText("موافق")
        button_box.button(QDialogButtonBox.Cancel).setText("إلغاء")
        
        layout.addWidget(button_box)

    def validate_and_accept(self):
        """التحقق من صحة البيانات قبل القبول"""
        if not self.name.text().strip():
            QMessageBox.warning(
                self,
                'خطأ في الإدخال',
                'يجب إدخال اسم السائق على الأقل!',
                QMessageBox.Ok
            )
            self.name.setFocus()  # تركيز على حقل الاسم
            return
        
        # إذا تم إدخال الاسم، نقبل النموذج
        self.accept()

    def get_driver_data(self):
        return {
            'name': self.name.text(),
            'address': self.address.text(),
            'age': self.age.text(),
            'national_id': self.national_id.text()
        }
class AddTruckDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('إضافة شاحنة جديدة')
        self.setModal(True)
        self.resize(400, 300)
        self.setupUI()
        
    def setupUI(self):
        layout = QVBoxLayout(self)
        form_group = QGroupBox("بيانات الشاحنة")
        form_layout = QGridLayout(form_group)
        
        self.truck_number = QLineEdit()
        form_layout.addWidget(QLabel("رقم الشاحنة:"), 0, 1)
        form_layout.addWidget(self.truck_number, 0, 0)
        
        self.truck_type = QLineEdit()
        form_layout.addWidget(QLabel("نوع الشاحنة:"), 1, 1)
        form_layout.addWidget(self.truck_type, 1, 0)
        
        self.model = QLineEdit()
        form_layout.addWidget(QLabel("موديل الشاحنة:"), 2, 1)
        form_layout.addWidget(self.model, 2, 0)
        
        self.ownership = QComboBox()
        self.ownership.addItems(["إيجار", "مملوكه لشركه HD"])
        form_layout.addWidget(QLabel("ملكية الشاحنة:"), 3, 1)
        form_layout.addWidget(self.ownership, 3, 0)
        
        layout.addWidget(form_group)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
    
    def accept(self):
        # التحقق من رقم الشاحنة
        if not self.truck_number.text().strip():
            QMessageBox.warning(self, "تنبيه", "يجب إدخال رقم الشاحنة")
            self.truck_number.setFocus()
            return
            
        # التحقق من نوع الشاحنة
        if not self.truck_type.text().strip():
            QMessageBox.warning(self, "تنبيه", "يجب إدخال نوع الشاحنة")
            self.truck_type.setFocus()
            return
            
        # التحقق من موديل الشاحنة
        if not self.model.text().strip():
            QMessageBox.warning(self, "تنبيه", "يجب إدخال موديل الشاحنة")
            self.model.setFocus()
            return
        
        super().accept()
        
    def get_truck_data(self):
        return {
            'truck_number': self.truck_number.text().strip(),
            'truck_type': self.truck_type.text().strip(),
            'model': self.model.text().strip(),
            'ownership': self.ownership.currentText()
        }

class AddCompanyDialog(QDialog):
    def __init__(self, countries, parent=None):
        super().__init__(parent)
        self.countries = countries
        self.setWindowTitle('إضافة شركة جديدة')
        self.setModal(True)
        self.resize(400, 300)
        self.setupUI()

    def accept(self):
        # التحقق من البيانات المطلوبة قبل الإضافة
        if not self.company_name.text().strip():
            QMessageBox.warning(self, "تنبيه", "يجب إدخال اسم الشركة")
            self.company_name.setFocus()
            return
            
        if not self.client_name.text().strip():
            QMessageBox.warning(self, "تنبيه", "يجب إدخال اسم العميل")
            self.client_name.setFocus()
            return
            
        if not self.address.text().strip():
            QMessageBox.warning(self, "تنبيه", "يجب إدخال عنوان الشركة")
            self.address.setFocus()
            return
            
        super().accept()        
        
    def setupUI(self):
        layout = QVBoxLayout(self)
        form_group = QGroupBox("بيانات الشركة")
        form_layout = QGridLayout(form_group)
        
        self.company_name = QLineEdit()
        form_layout.addWidget(QLabel("اسم الشركة:"), 0, 1)
        form_layout.addWidget(self.company_name, 0, 0)
        
        self.country = QComboBox()
        self.country.addItems(self.countries)
        form_layout.addWidget(QLabel("دولة الشركة:"), 1, 1)
        form_layout.addWidget(self.country, 1, 0)
        
        self.client_name = QLineEdit()
        form_layout.addWidget(QLabel("اسم العميل:"), 2, 1)
        form_layout.addWidget(self.client_name, 2, 0)
        
        self.deal_type = QComboBox()
        self.deal_type.addItems(["استيراد", "تصدير", "استيراد وتصدير"])
        form_layout.addWidget(QLabel("نوع المعامله:"), 3, 1)
        form_layout.addWidget(self.deal_type, 3, 0)
        
        self.address = QLineEdit()
        form_layout.addWidget(QLabel("عنوان الشركة:"), 4, 1)
        form_layout.addWidget(self.address, 4, 0)
        
        layout.addWidget(form_group)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        
    def get_company_data(self):
        try:
            return {
                'company_name': self.company_name.text(),
                'country': self.country.currentText(),
                'client_name': self.client_name.text(),
                'deal_type': self.deal_type.currentText(),
                'address': self.address.text()
            }
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ في معالجة البيانات: {str(e)}")
            return None

class AddExpenseDialog(QDialog):
    def __init__(self, trips, parent=None):
        super().__init__(parent)
        self.trips = trips
        self.setWindowTitle('إضافة مصاريف جديدة')
        self.setModal(True)
        self.resize(400, 500)
        self.setupUI()
        
    def setupUI(self):
        layout = QVBoxLayout(self)
        form_group = QGroupBox("بيانات المصاريف")
        form_layout = QGridLayout(form_group)
        
        self.shipment_number = QComboBox()
        self.shipment_number.addItems([trip['shipment_number'] for trip in self.trips])
        form_layout.addWidget(QLabel("رقم الشحنة:"), 0, 1)
        form_layout.addWidget(self.shipment_number, 0, 0)
        
        self.fuel_cost = QLineEdit()
        self.fuel_cost.setValidator(QDoubleValidator(0.0, 1000000.0, 2))
        form_layout.addWidget(QLabel("رسوم الوقود:"), 1, 1)
        form_layout.addWidget(self.fuel_cost, 1, 0)
        
        self.oil_cost = QLineEdit()
        self.oil_cost.setValidator(QDoubleValidator(0.0, 1000000.0, 2))
        form_layout.addWidget(QLabel("رسوم الزيت:"), 2, 1)
        form_layout.addWidget(self.oil_cost, 2, 0)
        
        self.maintenance_cost = QLineEdit()
        self.maintenance_cost.setValidator(QDoubleValidator(0.0, 1000000.0, 2))
        form_layout.addWidget(QLabel("مصاريف الصيانة:"), 3, 1)
        form_layout.addWidget(self.maintenance_cost, 3, 0)
        
        self.army_card_cost = QLineEdit()
        self.army_card_cost.setValidator(QDoubleValidator(0.0, 1000000.0, 2))
        form_layout.addWidget(QLabel("رسوم كارتة الجيش:"), 4, 1)
        form_layout.addWidget(self.army_card_cost, 4, 0)
        
        self.rental_cost = QLineEdit()
        self.rental_cost.setValidator(QDoubleValidator(0.0, 1000000.0, 2))
        form_layout.addWidget(QLabel("رسوم إيجار الشاحنة:"), 5, 1)
        form_layout.addWidget(self.rental_cost, 5, 0)
        
        self.driver_salary = QLineEdit()
        self.driver_salary.setValidator(QDoubleValidator(0.0, 1000000.0, 2))
        form_layout.addWidget(QLabel("راتب السائق:"), 6, 1)
        form_layout.addWidget(self.driver_salary, 6, 0)
        
        self.delay_fine = QLineEdit()
        self.delay_fine.setValidator(QDoubleValidator(0.0, 1000000.0, 2))
        form_layout.addWidget(QLabel("غرامة التأخير:"), 7, 1)
        form_layout.addWidget(self.delay_fine, 7, 0)
        
        self.transport_cost = QLineEdit()
        self.transport_cost.setValidator(QDoubleValidator(0.0, 1000000.0, 2))
        form_layout.addWidget(QLabel("تكلفة نقل الشحنة:"), 8, 1)
        form_layout.addWidget(self.transport_cost, 8, 0)
        
        layout.addWidget(form_group)
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.validate_and_accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        
    def validate_and_accept(self):
        expense_data = self.get_expense_data()
        if any(value > 0.0 for key, value in expense_data.items() if key != 'shipment_number'):
            self.accept()
        else:
            QMessageBox.warning(self, 'خطأ', 'يرجى إدخال قيمة واحدة على الأقل للمصاريف!')
        
    def get_expense_data(self):
        def safe_float(value):
            try:
                return float(value) if value.strip() else 0.0
            except ValueError:
                return 0.0
                
        return {
            'shipment_number': self.shipment_number.currentText(),
            'fuel_cost': safe_float(self.fuel_cost.text()),
            'oil_cost': safe_float(self.oil_cost.text()),
            'maintenance_cost': safe_float(self.maintenance_cost.text()),
            'army_card_cost': safe_float(self.army_card_cost.text()),
            'rental_cost': safe_float(self.rental_cost.text()),
            'driver_salary': safe_float(self.driver_salary.text()),
            'delay_fine': safe_float(self.delay_fine.text()),
            'transport_cost': safe_float(self.transport_cost.text())
        }

class AddTripDialog(QDialog):
    def __init__(self, countries, drivers_data, parent=None):
        super().__init__(parent)
        self.countries = countries
        self.drivers_data = drivers_data
        self.parent_app = parent  # حفظ المرجع للتطبيق الرئيسي
        self.setWindowTitle('إضافة شحنة جديدة')
        self.setModal(True)
        self.resize(500, 600)
        self.setupUI()
        
    def generate_shipment_number(self):
        letter = random.choice(string.ascii_uppercase)
        numbers = ''.join(random.choice(string.digits) for _ in range(4))
        return f"{letter}{numbers}"
        
    def setupUI(self):
        layout = QVBoxLayout(self)
        basic_group = QGroupBox("بيانات الشحنة الأساسية")
        basic_layout = QGridLayout(basic_group)
        
        self.shipment_number = QLineEdit()
        self.shipment_number.setText(self.generate_shipment_number())
        self.shipment_number.setReadOnly(True)
        basic_layout.addWidget(QLabel("رقم الشحنة:"), 0, 1)
        basic_layout.addWidget(self.shipment_number, 0, 0)
        
        self.driver_name = QComboBox()
        self.driver_name.addItems([driver['name'] for driver in self.drivers_data])
        self.driver_name.currentTextChanged.connect(self.check_driver_permissions)  # إضافة اتصال للإشارة
        basic_layout.addWidget(QLabel("اسم السائق:"), 1, 1)
        basic_layout.addWidget(self.driver_name, 1, 0)

        self.warning_label = QLabel()
        self.warning_label.setStyleSheet("color: red; font-weight: bold;")
        layout.insertWidget(2, self.warning_label)  # إضافته بعد الأدوات
        
        self.shipment_type = QComboBox()
        self.shipment_type.addItems(["استيراد", "تصدير",'استيراد/تصدير'])
        basic_layout.addWidget(QLabel("نوع الشحنة:"), 2, 1)
        basic_layout.addWidget(self.shipment_type, 2, 0)
        
        self.shipment_date = QDateEdit()
        self.shipment_date.setDate(QDate.currentDate())
        self.shipment_date.setCalendarPopup(True)
        basic_layout.addWidget(QLabel("تاريخ الشحنة:"), 3, 1)
        basic_layout.addWidget(self.shipment_date, 3, 0)
        
        self.allowance_period = QLineEdit()
        self.allowance_period.setText("2")
        self.allowance_period.setValidator(QIntValidator(1, 30))
        basic_layout.addWidget(QLabel("فترة السماح (أيام):"), 4, 1)
        basic_layout.addWidget(self.allowance_period, 4, 0)
        
        layout.addWidget(basic_group)
        
        destination_group = QGroupBox("بيانات الوجهات")
        dest_layout = QGridLayout(destination_group)
        
        self.start = QComboBox()
        self.start.addItems(self.countries)
        dest_layout.addWidget(QLabel("من:"), 0, 1)
        dest_layout.addWidget(self.start, 0, 0)
        
        self.final_destination = QComboBox()
        self.final_destination.addItems(self.countries)
        dest_layout.addWidget(QLabel("الوجهة النهائية:"), 1, 1)
        dest_layout.addWidget(self.final_destination, 1, 0)
        
        self.intermediate_country = QComboBox()
        self.intermediate_country.addItem("لا يوجد")
        self.intermediate_country.addItems(self.countries)
        dest_layout.addWidget(QLabel("الدولة الوسيطة:"), 2, 1)
        dest_layout.addWidget(self.intermediate_country, 2, 0)
        
        layout.addWidget(destination_group)
        
        status_group = QGroupBox("حالة الشحنة")
        status_layout = QGridLayout(status_group)
        
        self.shipment_status = QComboBox()
        self.shipment_status.addItems([
            "في الجمارك",
            "في جمارك الدولة الوسيطة", 
            "في الميناء",
            "في ميناء الدولة الوسيطة",
            "في جمارك الوجهة النهائية",
            "في ميناء الوجهة النهائية",
            "اكتملت وتم استلام الشحنة",
            "متأخرة"
        ])
        status_layout.addWidget(QLabel("حالة الشحنة:"), 0, 1)
        status_layout.addWidget(self.shipment_status, 0, 0)
        
        layout.addWidget(status_group)
        
        notes_group = QGroupBox("ملاحظات")
        notes_layout = QVBoxLayout(notes_group)
        self.notes = QTextEdit()
        self.notes.setMaximumHeight(80)
        notes_layout.addWidget(self.notes)
        layout.addWidget(notes_group)
    
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept)
        button_box.rejected.connect(self.reject)
        layout.addWidget(button_box)
        
    def check_driver_permissions(self):
        """التحقق من صلاحيات السائق للسفر إلى الدول المحددة"""
        driver_name = self.driver_name.currentText()
        final_destination = self.final_destination.currentText()
        intermediate_country = self.intermediate_country.currentText()
        
        # الحصول على بيانات السائق
        driver = next((d for d in self.drivers_data if d['name'] == driver_name), None)
        if not driver:
            return
            
        # الحصول على قائمة الدول المسموح للسائق بالسفر إليها
        allowed_countries = []
        for country in self.parent_app.countries_data:  # سنقوم بتعديل ShippingManagementApp لاحتواء countries_data
            if driver_name in country['allowed_drivers']:
                allowed_countries.append(country['name'])
        
        # التحقق من الوجهة النهائية
        if final_destination != "لا يوجد" and final_destination not in allowed_countries:
            self.warning_label.setText(f"تحذير: هذا السائق غير مسموح له بالسفر إلى {final_destination}")
            return
            
        # التحقق من الدولة الوسيطة إذا كانت محددة
        if intermediate_country != "لا يوجد" and intermediate_country not in allowed_countries:
            self.warning_label.setText(f"تحذير: هذا السائق غير مسموح له بالسفر إلى {intermediate_country}")
            return
            
        # إذا كل شيء مسموح
        self.warning_label.setText("")
    
    def get_trip_data(self):
        """تعديل الدالة للتحقق من الصلاحيات قبل القبول"""
        driver_name = self.driver_name.currentText()
        final_destination = self.final_destination.currentText()
        intermediate_country = self.intermediate_country.currentText() if self.intermediate_country.currentText() != "لا يوجد" else ""
        
        # الحصول على بيانات السائق
        driver = next((d for d in self.drivers_data if d['name'] == driver_name), None)
        if not driver:
            QMessageBox.warning(self, 'خطأ', 'السائق المحدد غير موجود!')
            return None
            
        # التحقق من الصلاحيات
        allowed_countries = []
        for country in self.parent_app.countries_data:
            if driver_name in country['allowed_drivers']:
                allowed_countries.append(country['name'])
        
        if final_destination not in allowed_countries:
            QMessageBox.warning(self, 'خطأ', f'هذا السائق غير مسموح له بالسفر إلى {final_destination}!')
            return None
            
        if intermediate_country and intermediate_country not in allowed_countries:
            QMessageBox.warning(self, 'خطأ', f'هذا السائق غير مسموح له بالسفر إلى {intermediate_country}!')
            return None
            
        # إذا كل شيء صحيح، إرجاع بيانات الشحنة
        return {
            'shipment_number': self.shipment_number.text(),
            'driver_name': driver_name,
            'shipment_type': self.shipment_type.currentText(),
            'shipment_date': self.shipment_date.date().toString('yyyy-MM-dd'),
            'start': self.start.currentText(),
            'final_destination': final_destination,
            'intermediate_country': intermediate_country,
            'shipment_status': self.shipment_status.currentText(),
            'allowance_period': self.allowance_period.text(),
            'notes': self.notes.toPlainText()
        }
    



class ShipmentGalleryDialog(QDialog):
    def __init__(self, trips_data, parent=None):
        super().__init__(parent)
        self.trips_data = trips_data
        self.parent_app = parent

        self.db_file = {
             "host": "localhost",
            "user": "hany",
            "password": "hany",
            "database": "transport_db"
        } 

        self.image_paths = self.load_image_data()
        self.setWindowTitle('معرض الشحنات')
        self.setModal(True)
        self.resize(800, 600)
        self.setupUI()
        
    def setupUI(self):
        layout = QVBoxLayout(self)
        
        # شريط الأدوات
        toolbar = QHBoxLayout()
        
        self.shipment_combo = QComboBox()
        self.shipment_combo.addItems([trip['shipment_number'] for trip in self.trips_data])
        self.shipment_combo.currentTextChanged.connect(self.load_shipment_images)
        toolbar.addWidget(QLabel("رقم الشحنة:"))
        toolbar.addWidget(self.shipment_combo)
        
        add_image_btn = QPushButton("إضافة صورة")
        add_image_btn.clicked.connect(self.add_image)
        toolbar.addWidget(add_image_btn)
        
        delete_image_btn = QPushButton("حذف الصورة المحددة")
        delete_image_btn.clicked.connect(self.delete_image)
        toolbar.addWidget(delete_image_btn)
        
        toolbar.addStretch()
        layout.addLayout(toolbar)
        
        # قائمة الصور
        self.images_list = QListWidget()
        self.images_list.itemDoubleClicked.connect(self.view_image)
        layout.addWidget(self.images_list)
        
        # زر الإغلاق
        button_box = QDialogButtonBox(QDialogButtonBox.Ok)
        button_box.accepted.connect(self.accept)
        layout.addWidget(button_box)
        
        # تحميل الصور للشحنة الأولى
        if self.trips_data:
            self.load_shipment_images(self.trips_data[0]['shipment_number'])
    
    def load_shipment_images(self, shipment_number):
        """تحميل الصور للشحنة المحددة"""
        self.images_list.clear()
        if shipment_number in self.image_paths:
            for image_path in self.image_paths[shipment_number]:
                item = QListWidgetItem(image_path)
                self.images_list.addItem(item)
    
    def view_image(self, item):
        """عرض الصورة المحددة"""
        from PyQt5.QtWidgets import QMessageBox
        from PyQt5.QtGui import QPixmap
        
        try:
            pixmap = QPixmap(item.text())
            if pixmap.isNull():
                QMessageBox.warning(self, "خطأ", "تعذر تحميل الصورة!")
                return
                
            # إنشاء نافذة لعرض الصورة
            image_dialog = QDialog(self)
            image_dialog.setWindowTitle("عرض الصورة")
            layout = QVBoxLayout(image_dialog)
            
            label = QLabel()
            label.setPixmap(pixmap.scaled(800, 600, Qt.KeepAspectRatio))
            layout.addWidget(label)
            
            button_box = QDialogButtonBox(QDialogButtonBox.Ok)
            button_box.accepted.connect(image_dialog.accept)
            layout.addWidget(button_box)
            
            image_dialog.exec_()
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ أثناء عرض الصورة: {str(e)}")
    def load_image_data(self):
            """تحميل بيانات الصور من قاعدة البيانات"""
            images_dict = {}
            try:
                conn = pymysql.connect(**self.db_file)
                cursor = conn.cursor()
                cursor.execute("SELECT shipment_number, image_path FROM shipment_images")
                
                for shipment_number, image_path in cursor.fetchall():
                    if shipment_number not in images_dict:
                        images_dict[shipment_number] = []
                    if os.path.exists(image_path):  # التحقق من وجود الصورة
                        images_dict[shipment_number].append(image_path)
                
                conn.close()
            except Exception as e:
                print(f"Error loading images: {str(e)}")
                QMessageBox.warning(self, 'خطأ', f'حدث خطأ أثناء تحميل الصور: {str(e)}')
            
            return images_dict

    def save_image_data(self):
        """حفظ بيانات الصور في قاعدة البيانات"""
        try:
            conn = pymysql.connect(**self.db_file)
            cursor = conn.cursor()
            
            # حذف جميع الصور القديمة
            cursor.execute("DELETE FROM shipment_images")
            
            # إدراج الصور الجديدة
            for shipment_number, paths in self.image_paths.items():
                for path in paths:
                    cursor.execute(
                        "INSERT INTO shipment_images (shipment_number, image_path) VALUES (%s, %s)",
                        (shipment_number, path)
                    )
            
            conn.commit()
            conn.close()
            
        except Exception as e:
            print(f"Error saving images: {str(e)}")
            QMessageBox.critical(self, 'خطأ', f'حدث خطأ أثناء حفظ الصور: {str(e)}')
    
   

    def add_image(self):
        """إضافة صورة جديدة للشحنة"""
        shipment_number = self.shipment_combo.currentText()
        try:
            # فتح نافذة اختيار الملف
            file_path, _ = QFileDialog.getOpenFileName(
                self, 
                "اختر صورة للشحنة", 
                "", 
                "ملفات الصور (*.png *.jpg *.jpeg *.bmp)"
            )
            
            if file_path:
                # نسخ الصورة إلى مجلد الصور الخاص بالتطبيق
                app_images_dir = os.path.join(os.path.dirname(self.db_file), "shipment_images")
                os.makedirs(app_images_dir, exist_ok=True)
                
                # إنشاء اسم فريد للصورة
                file_ext = os.path.splitext(file_path)[1]
                new_filename = f"{shipment_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{file_ext}"
                new_path = os.path.join(app_images_dir, new_filename)
                
                # نسخ الصورة
                import shutil
                shutil.copy2(file_path, new_path)
                
                # إضافة مسار الصورة الجديد
                if shipment_number not in self.image_paths:
                    self.image_paths[shipment_number] = []
                self.image_paths[shipment_number].append(new_path)
                
                # تحديث قاعدة البيانات
                self.save_image_data()
                
                # تحديث العرض
                self.load_shipment_images(shipment_number)
                
        except Exception as e:
            print(f"Error adding image: {str(e)}")
            QMessageBox.critical(self, 'خطأ', f'حدث خطأ أثناء إضافة الصورة: {str(e)}')

    def delete_image(self):
        """حذف الصورة المحددة"""
        try:
            current_item = self.images_list.currentItem()
            if current_item:
                shipment_number = self.shipment_combo.currentText()
                image_path = current_item.text()
                
                reply = QMessageBox.question(
                    self, 
                    'تأكيد الحذف',
                    'هل أنت متأكد من حذف هذه الصورة؟',
                    QMessageBox.Yes | QMessageBox.No
                )
                
                if reply == QMessageBox.Yes:
                    # حذف الملف
                    if os.path.exists(image_path):
                        os.remove(image_path)
                    
                    # حذف من القائمة
                    self.image_paths[shipment_number].remove(image_path)
                    if not self.image_paths[shipment_number]:
                        del self.image_paths[shipment_number]
                    
                    # تحديث قاعدة البيانات
                    self.save_image_data()
                    
                    # تحديث العرض
                    self.load_shipment_images(shipment_number)
                    
        except Exception as e:
            print(f"Error deleting image: {str(e)}")
            QMessageBox.critical(self, 'خطأ', f'حدث خطأ أثناء حذف الصورة: {str(e)}')

    def accept(self):
        """حفظ التغييرات عند إغلاق النافذة"""
        self.save_image_data()
        super().accept()

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class ShippingManagementApp(QMainWindow):
    def __init__(self):
        super().__init__()
        icon_path = resource_path('logo.ico')
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

        self.db_file = {
             "host": "localhost",
            "user": "hany",
            "password": "hany",
            "database": "transport_db"
        } 
        self.images_dir = resource_path("shipment_images")
        os.makedirs(self.images_dir, exist_ok=True)
        
        self.init_database()
        self.current_user = None
        self.countries_data = self.load_countries_data()
        self.countries = [c['name'] for c in self.countries_data]
        self.trips_data = self.load_trips_data()
        self.drivers_data = self.load_drivers_data()
        self.trucks_data = self.load_trucks_data()
        self.companies_data = self.load_companies_data()
        self.expenses_data = self.load_expenses_data()
        self.show_login()

    def init_database(self):
        """Initialize the SQLite database and create tables if they don't exist."""
        self.db_file = {
        "host": "localhost",
        "user": "hany",
        "password": "hany",
        "database": "transport_db"
    }
        conn = pymysql.connect(**self.db_file)
        

    def load_users(self):
        """Load users from the database."""
        conn = pymysql.connect(**self.db_file)
        cursor = conn.cursor()
        cursor.execute("SELECT username, password, type, permissions FROM users")
        users = [
            {
                'username': row[0],
                'password': row[1],
                'type': row[2],
                'permissions': json.loads(row[3])
            }
            for row in cursor.fetchall()
        ]
        conn.close()
        return users

    def load_countries_data(self):
        """تحميل بيانات الدول من قاعدة البيانات"""
        try:
            print("جاري تحميل بيانات الدول من قاعدة البيانات...")  # رسالة debug
            conn = pymysql.connect(**self.db_file)
            cursor = conn.cursor()
            
            # التحقق من وجود الجدول
            cursor.execute("SHOW TABLES LIKE %s", ("countries",))
            table_exists = cursor.fetchone()
            
            cursor.execute("SELECT name, allowed_drivers FROM countries")
            countries = []
            
            for row in cursor.fetchall():
                try:
                    countries.append({
                        'name': row[0],
                        'allowed_drivers': json.loads(row[1]) if row[1] else []
                    })
                except json.JSONDecodeError:
                    print(f"خطأ في تحويل JSON للسائقين في دولة {row[0]}")  # رسالة debug
                    countries.append({
                        'name': row[0],
                        'allowed_drivers': []
                    })
            
            conn.close()
            print(f"تم تحميل {len(countries)} دولة بنجاح")  # رسالة debug
            return countries
            
        except sqlite3.Error as e:
            print(f"خطأ في sqlite أثناء تحميل الدول: {str(e)}")  # رسالة debug
            QMessageBox.critical(self, 'خطأ', f'حدث خطأ في قاعدة البيانات: {str(e)}')
            return []
            
        except Exception as e:
            print(f"خطأ غير متوقع أثناء تحميل الدول: {str(e)}")  # رسالة debug
            QMessageBox.critical(self, 'خطأ', f'حدث خطأ غير متوقع: {str(e)}')
            return []

    def load_drivers_data(self):
        """Load drivers data from the database."""
        conn = pymysql.connect(**self.db_file)
        cursor = conn.cursor()
        cursor.execute("SELECT name, address, age, national_id FROM drivers")
        drivers = [
            {
                'name': row[0],
                'address': row[1],
                'age': str(row[2]) if row[2] is not None else '',
                'national_id': row[3]
            }
            for row in cursor.fetchall()
        ]
        conn.close()
        return drivers

    def load_trucks_data(self):
        """Load trucks data from the database."""
        conn = pymysql.connect(**self.db_file)
        cursor = conn.cursor()
        cursor.execute("SELECT truck_number, truck_type, model, ownership FROM trucks")
        trucks = [
            {
                'truck_number': row[0],
                'truck_type': row[1],
                'model': row[2],
                'ownership': row[3]
            }
            for row in cursor.fetchall()
        ]
        conn.close()
        return trucks

    def load_companies_data(self):
        """Load companies data from the database."""
        conn = pymysql.connect(**self.db_file)
        cursor = conn.cursor()
        cursor.execute("SELECT company_name, country, client_name, deal_type, address FROM companies")
        companies = [
            {
                'company_name': row[0],
                'country': row[1],
                'client_name': row[2],
                'deal_type': row[3],
                'address': row[4]
            }
            for row in cursor.fetchall()
        ]
        conn.close()
        return companies

    def load_trips_data(self):
        """Load trips data from the database."""
        conn = pymysql.connect(**self.db_file)
        cursor = conn.cursor()
        cursor.execute("SELECT shipment_number, driver_name, shipment_type, shipment_date, start, final_destination, intermediate_country, shipment_status, allowance_period, notes FROM trips")
        trips = [
            {
                'shipment_number': row[0],
                'driver_name': row[1],
                'shipment_type': row[2],
                'shipment_date': row[3],
                'start': row[4],
                'final_destination': row[5],
                'intermediate_country': row[6],
                'shipment_status': row[7],
                'allowance_period': str(row[8]),
                'notes': row[9]
            }
            for row in cursor.fetchall()
        ]
        conn.close()
        return trips

    def load_expenses_data(self):
        """Load expenses data from the database."""
        conn = pymysql.connect(**self.db_file)
        cursor = conn.cursor()
        cursor.execute("SELECT shipment_number, fuel_cost, oil_cost, maintenance_cost, army_card_cost, rental_cost, driver_salary, delay_fine, transport_cost FROM expenses")
        expenses = [
            {
                'shipment_number': row[0],
                'fuel_cost': float(row[1]),
                'oil_cost': float(row[2]),
                'maintenance_cost': float(row[3]),
                'army_card_cost': float(row[4]),
                'rental_cost': float(row[5]),
                'driver_salary': float(row[6]),
                'delay_fine': float(row[7]),
                'transport_cost': float(row[8])
            }
            for row in cursor.fetchall()
        ]
        conn.close()
        return expenses
        
    
        
    def initUI(self):
        self.setWindowTitle('إدارة شركة النقل الثقيل')
        self.setGeometry(100, 100, 1400, 800)
        
        # ضبط اتجاه التطبيق بأكمله إلى اليمين إلى اليسار
        self.setLayoutDirection(Qt.RightToLeft)
        
        font = QFont("Arial", 12)
        self.setFont(font)
        
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        
        main_layout = QHBoxLayout(main_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        
        self.create_sidebar()
        main_layout.addWidget(self.sidebar)
        
        self.main_content = QWidget()
        self.main_content_layout = QVBoxLayout(self.main_content)
        main_layout.addWidget(self.main_content)
        
        if self.current_user['permissions']['trips']:
            self.show_trips()
        else:
            self.clear_main_content()
            self.main_content_layout.addWidget(QLabel("غير مسموح لك بالوصول إلى أي قسم!"))
    def show_gallery(self):
        """عرض معرض صور الشحنات"""
        if not hasattr(self, 'trips_data') or not self.trips_data:
            QMessageBox.information(self, 'تنبيه', 'لا توجد شحنات لعرضها!')
            return
            
        if not self.current_user['permissions'].get('gallery', False):
            QMessageBox.warning(self, 'خطأ', 'غير مسموح لك بالوصول إلى معرض الشحنات!')
            return
            
        dialog = ShipmentGalleryDialog(self.trips_data, self)
        dialog.exec_()
        
    def create_sidebar(self):
        self.sidebar = QFrame()
        self.sidebar.setFixedWidth(220)
        self.sidebar.setStyleSheet("""
            QFrame {
                background-color: #4a5568;
                color: white;
            }
        """)
        
        sidebar_layout = QVBoxLayout(self.sidebar)
        sidebar_layout.setContentsMargins(0, 0, 0, 0)
        sidebar_layout.setSpacing(0)
        
        title_label = QLabel("إدارة شركة\nالنقل الثقيل")
        title_label.setAlignment(Qt.AlignCenter)
        title_label.setStyleSheet("""
            QLabel {
                background-color: #2d3748;
                color: white;
                padding: 20px;
                font-size: 16px;
                font-weight: bold;
                border-bottom: 1px solid #4a5568;
            }
        """)
        sidebar_layout.addWidget(title_label)
        
        menu_buttons = [
            ("🚛 الشحنات", "trips_btn", self.show_trips, 'trips'),
            ("👥 السائقين", "drivers_btn", self.show_drivers, 'drivers'),
            ("🌍 إدارة الدول", "countries_btn", self.manage_countries, 'countries'),
            ("🚚 الشاحنات", "trucks_btn", self.show_trucks, 'trucks'),
            ("🏢 الشركات", "companies_btn", self.show_companies, 'companies'),
            ("💰 المصاريف", "expenses_btn", self.show_expenses, 'expenses'),
            ("📊 التقارير", "reports_btn", self.show_reports, 'reports'),
            ("🖼️ معرض الشحنات", "gallery_btn", self.show_gallery, 'gallery'),
            ("⚙️ الإعدادات", "settings_btn", self.show_settings, 'settings'),
            ("🚪 تسجيل الخروج", "logout_btn", self.logout, None)
        ]
        
        for btn_text, btn_name, callback, permission in menu_buttons:
            btn = QPushButton(btn_text)
            btn.setObjectName(btn_name)
            btn.clicked.connect(lambda checked, cb=callback, perm=permission: self.check_permission_and_show(cb, perm))
            btn.setStyleSheet("""
                QPushButton {
                    background-color: transparent;
                    color: white;
                    text-align: left;
                    padding: 15px 20px;
                    border: none;
                    font-size: 14px;
                }
                QPushButton:hover {
                    background-color: #2d3748;
                }
                QPushButton:pressed {
                    background-color: #1a202c;
                }
            """)
            sidebar_layout.addWidget(btn)
        
        sidebar_layout.addStretch()
    def check_permission_and_show(self, callback, permission):
        """دالة جديدة للتحقق من الصلاحيات وإظهار الرسائل"""
        if permission is None:  # للأزرار التي لا تحتاج صلاحية مثل تسجيل الخروج
            callback()
            return
        
        if permission == 'settings':
            if self.current_user['type'] == 'Admin':
                callback()
            else:
                QMessageBox.warning(self, 'خطأ', 'غير مسموح لك بالوصول إلى إعدادات المستخدمين!')
        else:
            if self.current_user['permissions'].get(permission, False):
                callback()
            else:
                QMessageBox.warning(self, 'خطأ', f'غير مسموح لك بالوصول إلى قسم {permission}!')
        
        
    def logout(self):
        try:
            print("Logging out...")  # للتتبع
            self.clear_main_content()
            self.current_user = None
            self.hide()
            
            # إعادة تهيئة البيانات
            self.countries_data = self.load_countries_data()
            self.countries = [c['name'] for c in self.countries_data]
            self.trips_data = self.load_trips_data()
            self.drivers_data = self.load_drivers_data()
            self.trucks_data = self.load_trucks_data()
            self.companies_data = self.load_companies_data()
            self.expenses_data = self.load_expenses_data()
            
            self.show_login()
            
        except Exception as e:
            print(f"Error during logout: {str(e)}")
            QMessageBox.critical(self, 'خطأ', f'حدث خطأ أثناء تسجيل الخروج: {str(e)}')

    def show_login(self):
        try:
            print("Showing login dialog...")  # للتتبع
            login_dialog = LoginDialog(self)
            result = login_dialog.exec_()
            
            if result == QDialog.Accepted:
                print("Login accepted, initializing UI...")  # للتتبع
                self.current_user = login_dialog.current_user
                # إعادة إنشاء واجهة المستخدم
                if self.centralWidget():
                    self.centralWidget().deleteLater()
                self.initUI()
                self.show()
            else:
                print("Login rejected, closing application...")  # للتتبع
                self.close()  # استخدام close() بدلاً من sys.exit()
                
        except Exception as e:
            print(f"Error during login: {str(e)}")
            QMessageBox.critical(self, 'خطأ', f'حدث خطأ أثناء تسجيل الدخول: {str(e)}')
            self.close()
        
    def create_trips_content(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        
        header_layout = QHBoxLayout()
        user_label = QLabel(f"{self.current_user['username']} ({self.current_user['type']})")
        user_label.setAlignment(Qt.AlignRight)
        user_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                color: #2d3748;
            }
        """)
        header_layout.addStretch()
        header_layout.addWidget(user_label)
        layout.addLayout(header_layout)
        
        toolbar_layout = QHBoxLayout()
        trips_title = QLabel("الشحنات الحالية")
        trips_title.setStyleSheet("""
            QLabel {
                font-size: 24px;
                font-weight: bold;
                color: #2d3748;
                margin: 20px 0;
            }
        """)
        toolbar_layout.addWidget(trips_title)
        toolbar_layout.addStretch()
        
        self.trips_filter = QComboBox()
        self.trips_filter.addItems([
            "عرض الجميع",
            "في الجمارك",
            "في جمارك الدولة الوسيطة",
            "في الميناء",
            "في ميناء الدولة الوسيطة",
            "في جمارك الوجهة النهائية",
            "في ميناء الوجهة النهائية",
            "اكتملت وتم استلام الشحنة",
            "متأخرة"
        ])
        self.trips_filter.currentTextChanged.connect(self.update_trips_table)
        toolbar_layout.addWidget(QLabel("التصنيف:"))
        toolbar_layout.addWidget(self.trips_filter)
        
        add_trip_btn = QPushButton("إضافة شحنة جديدة")
        add_trip_btn.clicked.connect(self.add_new_trip)
        add_trip_btn.setStyleSheet("""
            QPushButton {
                background-color: #4299e1;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #3182ce;
            }
        """)
        toolbar_layout.addWidget(add_trip_btn)
        
        edit_trip_btn = QPushButton("تعديل الشحنة")
        edit_trip_btn.clicked.connect(self.edit_trip)
        edit_trip_btn.setStyleSheet("""
            QPushButton {
                background-color: #ecc94b;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d69e2e;
            }
        """)
        toolbar_layout.addWidget(edit_trip_btn)
        
        delete_trip_btn = QPushButton("حذف الشحنة")
        delete_trip_btn.clicked.connect(self.delete_trip)
        delete_trip_btn.setStyleSheet("""
            QPushButton {
                background-color: #e53e3e;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c53030;
            }
        """)
        toolbar_layout.addWidget(delete_trip_btn)
        
        print_btn = QPushButton("طباعة")
        print_btn.clicked.connect(self.print_trips_table)
        print_btn.setStyleSheet("""
            QPushButton {
                background-color: #38a169;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2f855a;
            }
        """)
        toolbar_layout.addWidget(print_btn)
        
        export_btn = QPushButton("تصدير إلى Excel")
        export_btn.clicked.connect(self.export_trips_to_excel)
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #ecc94b;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d69e2e;
            }
        """)
        toolbar_layout.addWidget(export_btn)
        
        layout.addLayout(toolbar_layout)
        self.create_trips_table()
        layout.addWidget(self.trips_table)
        
        return widget
        
    def create_drivers_content(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        
        header_layout = QHBoxLayout()
        user_label = QLabel(f"{self.current_user['username']} ({self.current_user['type']})")
        user_label.setAlignment(Qt.AlignRight)
        user_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                color: #2d3748;
            }
        """)
        header_layout.addStretch()
        header_layout.addWidget(user_label)
        layout.addLayout(header_layout)
        
        toolbar_layout = QHBoxLayout()
        drivers_title = QLabel("السائقين الحاليين")
        drivers_title.setStyleSheet("""
            QLabel {
                font-size: 24px;
                font-weight: bold;
                color: #2d3748;
                margin: 20px 0;
            }
        """)
        toolbar_layout.addWidget(drivers_title)
        toolbar_layout.addStretch()
        
        add_driver_btn = QPushButton("إضافة سائق جديد")
        add_driver_btn.clicked.connect(self.add_new_driver)
        add_driver_btn.setStyleSheet("""
            QPushButton {
                background-color: #4299e1;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #3182ce;
            }
        """)
        toolbar_layout.addWidget(add_driver_btn)
        
        edit_driver_btn = QPushButton("تعديل السائق")
        edit_driver_btn.clicked.connect(self.edit_driver)
        edit_driver_btn.setStyleSheet("""
            QPushButton {
                background-color: #ecc94b;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d69e2e;
            }
        """)
        toolbar_layout.addWidget(edit_driver_btn)
        
        delete_driver_btn = QPushButton("حذف السائق")
        delete_driver_btn.clicked.connect(self.delete_driver)
        delete_driver_btn.setStyleSheet("""
            QPushButton {
                background-color: #e53e3e;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c53030;
            }
        """)
        toolbar_layout.addWidget(delete_driver_btn)
        
        print_btn = QPushButton("طباعة")
        print_btn.clicked.connect(self.print_drivers_table)
        print_btn.setStyleSheet("""
            QPushButton {
                background-color: #38a169;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2f855a;
            }
        """)
        toolbar_layout.addWidget(print_btn)
        
        export_btn = QPushButton("تصدير إلى Excel")
        export_btn.clicked.connect(self.export_drivers_to_excel)
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #ecc94b;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d69e2e;
            }
        """)
        toolbar_layout.addWidget(export_btn)
        
        layout.addLayout(toolbar_layout)
        self.create_drivers_table()
        layout.addWidget(self.drivers_table)
        
        return widget
        
    def create_trucks_content(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        
        header_layout = QHBoxLayout()
        user_label = QLabel(f"{self.current_user['username']} ({self.current_user['type']})")
        user_label.setAlignment(Qt.AlignRight)
        user_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                color: #2d3748;
            }
        """)
        header_layout.addStretch()
        header_layout.addWidget(user_label)
        layout.addLayout(header_layout)
        
        toolbar_layout = QHBoxLayout()
        trucks_title = QLabel("الشاحنات الحالية")
        trucks_title.setStyleSheet("""
            QLabel {
                font-size: 24px;
                font-weight: bold;
                color: #2d3748;
                margin: 20px 0;
            }
        """)
        toolbar_layout.addWidget(trucks_title)
        toolbar_layout.addStretch()
        
        self.trucks_filter = QComboBox()
        self.trucks_filter.addItems(["عرض الجميع", "إيجار", "مملوكه لشركه HD"])
        self.trucks_filter.currentTextChanged.connect(self.update_trucks_table)
        toolbar_layout.addWidget(QLabel("التصنيف:"))
        toolbar_layout.addWidget(self.trucks_filter)
        
        add_truck_btn = QPushButton("إضافة شاحنة جديدة")
        add_truck_btn.clicked.connect(self.add_new_truck)
        add_truck_btn.setStyleSheet("""
            QPushButton {
                background-color: #4299e1;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #3182ce;
            }
        """)
        toolbar_layout.addWidget(add_truck_btn)
        
        edit_truck_btn = QPushButton("تعديل الشاحنة")
        edit_truck_btn.clicked.connect(self.edit_truck)
        edit_truck_btn.setStyleSheet("""
            QPushButton {
                background-color: #ecc94b;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d69e2e;
            }
        """)
        toolbar_layout.addWidget(edit_truck_btn)
        
        delete_truck_btn = QPushButton("حذف الشاحنة")
        delete_truck_btn.clicked.connect(self.delete_truck)
        delete_truck_btn.setStyleSheet("""
            QPushButton {
                background-color: #e53e3e;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c53030;
            }
        """)
        toolbar_layout.addWidget(delete_truck_btn)
        
        print_btn = QPushButton("طباعة")
        print_btn.clicked.connect(self.print_trucks_table)
        print_btn.setStyleSheet("""
            QPushButton {
                background-color: #38a169;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2f855a;
            }
        """)
        toolbar_layout.addWidget(print_btn)
        
        export_btn = QPushButton("تصدير إلى Excel")
        export_btn.clicked.connect(self.export_trucks_to_excel)
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #ecc94b;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d69e2e;
            }
        """)
        toolbar_layout.addWidget(export_btn)
        
        layout.addLayout(toolbar_layout)
        self.create_trucks_table()
        layout.addWidget(self.trucks_table)
        
        return widget
        
    def create_companies_content(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)
        
        header_layout = QHBoxLayout()
        user_label = QLabel(f"{self.current_user['username']} ({self.current_user['type']})")
        user_label.setAlignment(Qt.AlignRight)
        user_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                color: #2d3748;
            }
        """)
        header_layout.addStretch()
        header_layout.addWidget(user_label)
        layout.addLayout(header_layout)
        
        toolbar_layout = QHBoxLayout()
        companies_title = QLabel("الشركات الحالية")
        companies_title.setStyleSheet("""
            QLabel {
                font-size: 24px;
                font-weight: bold;
                color: #2d3748;
                margin: 20px 0;
            }
        """)
        toolbar_layout.addWidget(companies_title)
        toolbar_layout.addStretch()
        
        add_company_btn = QPushButton("إضافة شركة جديدة")
        add_company_btn.clicked.connect(self.add_new_company)
        add_company_btn.setStyleSheet("""
            QPushButton {
                background-color: #4299e1;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #3182ce;
            }
        """)
        toolbar_layout.addWidget(add_company_btn)
        
        edit_company_btn = QPushButton("تعديل الشركة")
        edit_company_btn.clicked.connect(self.edit_company)
        edit_company_btn.setStyleSheet("""
            QPushButton {
                background-color: #ecc94b;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d69e2e;
            }
        """)
        toolbar_layout.addWidget(edit_company_btn)
        
        delete_company_btn = QPushButton("حذف الشركة")
        delete_company_btn.clicked.connect(self.delete_company)
        delete_company_btn.setStyleSheet("""
            QPushButton {
                background-color: #e53e3e;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c53030;
            }
        """)
        toolbar_layout.addWidget(delete_company_btn)
        
        print_btn = QPushButton("طباعة")
        print_btn.clicked.connect(self.print_companies_table)
        print_btn.setStyleSheet("""
            QPushButton {
                background-color: #38a169;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2f855a;
            }
        """)
        toolbar_layout.addWidget(print_btn)
        
        export_btn = QPushButton("تصدير إلى Excel")
        export_btn.clicked.connect(self.export_companies_to_excel)
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #ecc94b;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d69e2e;
            }
        """)
        toolbar_layout.addWidget(export_btn)
        
        layout.addLayout(toolbar_layout)
        self.create_companies_table()
        layout.addWidget(self.companies_table)
        
        return widget
        
    def create_expenses_content(self):
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)

        # Header section with user info
        header_layout = QHBoxLayout()
        user_label = QLabel(f"{self.current_user['username']} ({self.current_user['type']})")
        user_label.setAlignment(Qt.AlignRight)
        user_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                color: #2d3748;
            }
        """)
        header_layout.addStretch()
        header_layout.addWidget(user_label)
        layout.addLayout(header_layout)
        
        # Toolbar section
        toolbar_layout = QHBoxLayout()
        expenses_title = QLabel("المصاريف الحالية")
        expenses_title.setStyleSheet("""
            QLabel {
                font-size: 24px;
                font-weight: bold;
                color: #2d3748;
                margin: 20px 0;
            }
        """)
        toolbar_layout.addWidget(expenses_title)
        toolbar_layout.addStretch()
        
        add_expense_btn = QPushButton("إضافة مصاريف جديدة")
        add_expense_btn.clicked.connect(self.add_new_expense)
        add_expense_btn.setStyleSheet("""
            QPushButton {
                background-color: #4299e1;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #3182ce;
            }
        """)
        toolbar_layout.addWidget(add_expense_btn)
        
        edit_expense_btn = QPushButton("تعديل المصاريف")
        edit_expense_btn.clicked.connect(self.edit_expense)
        edit_expense_btn.setStyleSheet("""
            QPushButton {
                background-color: #ecc94b;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d69e2e;
            }
        """)
        toolbar_layout.addWidget(edit_expense_btn)
        
        delete_expense_btn = QPushButton("حذف المصاريف")
        delete_expense_btn.clicked.connect(self.delete_expense)
        delete_expense_btn.setStyleSheet("""
            QPushButton {
                background-color: #e53e3e;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #c53030;
            }
        """)
        toolbar_layout.addWidget(delete_expense_btn)
        
        print_btn = QPushButton("طباعة")
        print_btn.clicked.connect(self.print_expenses_table)
        print_btn.setStyleSheet("""
            QPushButton {
                background-color: #38a169;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2f855a;
            }
        """)
        toolbar_layout.addWidget(print_btn)
        
        export_btn = QPushButton("تصدير إلى Excel")
        export_btn.clicked.connect(self.export_expenses_to_excel)
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #ecc94b;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d69e2e;
            }
        """)
        toolbar_layout.addWidget(export_btn)
        
        layout.addLayout(toolbar_layout)

        # Create or reuse the expenses table - ONLY ONCE
        if not hasattr(self, 'expenses_table') or self.expenses_table is None:
            self.expenses_table = QTableWidget()
            self.expenses_table.setColumnCount(9)
            self.expenses_table.setHorizontalHeaderLabels([
                "رقم الشحنة", "راتب السائق", "رسوم الوقود", "رسوم الزيت",
                "مصاريف الصيانة", "رسوم كارتة الجيش", "رسوم إيجار الشاحنة",
                "غرامة التأخير", "تكلفة نقل الشحنة"
            ])
            self.expenses_table.setEditTriggers(QTableWidget.NoEditTriggers)
            self.expenses_table.horizontalHeader().setStretchLastSection(True)
            self.expenses_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            self.expenses_table.setAlternatingRowColors(True)
            self.expenses_table.setSelectionBehavior(QTableWidget.SelectRows)
        
        # Update the table with current data
        self.update_expenses_table()
        
        # Add the table to layout
        layout.addWidget(self.expenses_table)
        
        return widget
    def create_trips_table(self):
        self.trips_table = QTableWidget()
        self.trips_table.setColumnCount(10)  # تغيير من 9 إلى 10 أعمدة
        self.trips_table.setHorizontalHeaderLabels([
            "رقم الشحنة", "تاريخ الشحنة", "اسم السائق", "نوع الشحنة", 
            "مسافرة من", "الوجهة النهائية", "الدولة الوسيطة", "حالة الشحنة", 
            "فترة السماح", "ملاحظات"
        ])
        self.trips_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.update_trips_table()
        self.trips_table.horizontalHeader().setStretchLastSection(True)
        self.trips_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.trips_table.setAlternatingRowColors(True)
        self.trips_table.setSelectionBehavior(QTableWidget.SelectRows)
        
    def create_drivers_table(self):
        self.drivers_table = QTableWidget()
        self.drivers_table.setColumnCount(4)
        self.drivers_table.setHorizontalHeaderLabels([
            "الرقم القومي", "اسم السائق", "السن", "العنوان"
        ])
        self.drivers_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.update_drivers_table()
        self.drivers_table.horizontalHeader().setStretchLastSection(True)
        self.drivers_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.drivers_table.setAlternatingRowColors(True)
        self.drivers_table.setSelectionBehavior(QTableWidget.SelectRows)
        
    def create_trucks_table(self):
        self.trucks_table = QTableWidget()
        self.trucks_table.setColumnCount(4)
        self.trucks_table.setHorizontalHeaderLabels([
            "رقم الشاحنة", "نوع الشاحنة", "موديل الشاحنة", "ملكية الشاحنة"
        ])
        self.trucks_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.update_trucks_table()
        self.trucks_table.horizontalHeader().setStretchLastSection(True)
        self.trucks_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.trucks_table.setAlternatingRowColors(True)
        self.trucks_table.setSelectionBehavior(QTableWidget.SelectRows)
        
    def create_companies_table(self):
        self.companies_table = QTableWidget()
        self.companies_table.setColumnCount(5)
        self.companies_table.setHorizontalHeaderLabels([
            "اسم الشركة", "دولة الشركة", "اسم العميل", "نوع المعامله", "عنوان الشركة"
        ])
        self.companies_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.update_companies_table()
        self.companies_table.horizontalHeader().setStretchLastSection(True)
        self.companies_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.companies_table.setAlternatingRowColors(True)
        self.companies_table.setSelectionBehavior(QTableWidget.SelectRows)
     #h   
    def create_expenses_table(self):
        """This method now only updates the existing table instead of creating a new one"""
        if hasattr(self, 'expenses_table') and self.expenses_table is not None:
            self.update_expenses_table()
        else:
            # If for some reason the table doesn't exist, create it
            self.expenses_table = QTableWidget()
            self.expenses_table.setColumnCount(9)
            self.expenses_table.setHorizontalHeaderLabels([
                "رقم الشحنة", "راتب السائق", "رسوم الوقود", "رسوم الزيت",
                "مصاريف الصيانة", "رسوم كارتة الجيش", "رسوم إيجار الشاحنة",
                "غرامة التأخير", "تكلفة نقل الشحنة"
            ])
            self.expenses_table.setEditTriggers(QTableWidget.NoEditTriggers)
            self.expenses_table.horizontalHeader().setStretchLastSection(True)
            self.expenses_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            self.expenses_table.setAlternatingRowColors(True)
            self.expenses_table.setSelectionBehavior(QTableWidget.SelectRows)
            self.update_expenses_table()
        
    def create_reports_table(self):
        """Create or reuse the reports table"""
        if not hasattr(self, 'reports_table') or self.reports_table is None:
            self.reports_table = QTableWidget()
            self.reports_table.setColumnCount(4)
            self.reports_table.setHorizontalHeaderLabels([
                "رقم الشحنة", "إجمالي المصاريف", "الإيرادات", "صافي الربح"
            ])
            self.reports_table.setEditTriggers(QTableWidget.NoEditTriggers)
            self.reports_table.horizontalHeader().setStretchLastSection(True)
            self.reports_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            self.reports_table.setAlternatingRowColors(True)
            self.reports_table.setSelectionBehavior(QTableWidget.SelectRows)
        
        # Update the table with current data
        self.update_reports_table()
        
    def update_trips_table(self):
        filter_status = self.trips_filter.currentText() if hasattr(self, 'trips_filter') else "عرض الجميع"
        filtered_trips = [
            trip for trip in self.trips_data
            if filter_status == "عرض الجميع" or trip['shipment_status'] == filter_status
        ]
        
        self.trips_table.setRowCount(len(filtered_trips))
        current_date = datetime.now()
        
        for row, trip in enumerate(filtered_trips):
            shipment_date = datetime.strptime(trip['shipment_date'], '%Y-%m-%d')
            allowance_days = int(trip['allowance_period'])
            due_date = shipment_date + timedelta(days=allowance_days)
            
            is_delayed = current_date > due_date and trip['shipment_status'] != "اكتملت وتم استلام الشحنة"
            if is_delayed:
                trip['shipment_status'] = "متأخرة"
            
            self.trips_table.setItem(row, 0, QTableWidgetItem(trip['shipment_number']))
            self.trips_table.setItem(row, 1, QTableWidgetItem(trip['shipment_date']))
            self.trips_table.setItem(row, 2, QTableWidgetItem(trip['driver_name']))
            self.trips_table.setItem(row, 3, QTableWidgetItem(trip['shipment_type']))
            self.trips_table.setItem(row, 4, QTableWidgetItem(trip.get('start', '')))  # إضافة عمود "مسافرة من"
            self.trips_table.setItem(row, 5, QTableWidgetItem(trip['final_destination']))
            self.trips_table.setItem(row, 6, QTableWidgetItem(trip['intermediate_country'] or "لا يوجد"))
            
            status_item = QTableWidgetItem(trip['shipment_status'])
            status_item.setForeground(QColor(0, 0, 0))
            if trip['shipment_status'] == "اكتملت وتم استلام الشحنة":
                status_item.setBackground(QColor(255, 255, 224))
                status_item.setData(Qt.BackgroundRole, QColor(255, 255, 224))
            elif trip['shipment_status'] == "متأخرة":
                status_item.setBackground(QColor(240, 128, 128))
                status_item.setData(Qt.BackgroundRole, QColor(240, 128, 128))
            elif "جمارك" in trip['shipment_status']:
                status_item.setBackground(QColor(255, 255, 204))
                status_item.setData(Qt.BackgroundRole, QColor(255, 255, 204))
            elif "ميناء" in trip['shipment_status']:
                status_item.setBackground(QColor(204, 255, 255))
                status_item.setData(Qt.BackgroundRole, QColor(204, 255, 255))
                
            self.trips_table.setItem(row, 7, status_item)
            self.trips_table.setItem(row, 8, QTableWidgetItem(trip['allowance_period']))
            self.trips_table.setItem(row, 9, QTableWidgetItem(trip['notes'][:50] + "..." if len(trip['notes']) > 50 else trip['notes']))
    
    def update_drivers_table(self):
        self.drivers_table.setRowCount(len(self.drivers_data))
        
        for row, driver in enumerate(self.drivers_data):
            self.drivers_table.setItem(row, 0, QTableWidgetItem(driver['national_id']))
            self.drivers_table.setItem(row, 1, QTableWidgetItem(driver['name']))
            self.drivers_table.setItem(row, 2, QTableWidgetItem(driver['age']))
            self.drivers_table.setItem(row, 3, QTableWidgetItem(driver['address']))
    
    def update_trucks_table(self):
        filter_ownership = self.trucks_filter.currentText() if hasattr(self, 'trucks_filter') else "عرض الجميع"
        filtered_trucks = [
            truck for truck in self.trucks_data
            if filter_ownership == "عرض الجميع" or truck['ownership'] == filter_ownership
        ]
        
        self.trucks_table.setRowCount(len(filtered_trucks))
        
        for row, truck in enumerate(filtered_trucks):
            self.trucks_table.setItem(row, 0, QTableWidgetItem(truck['truck_number']))
            self.trucks_table.setItem(row, 1, QTableWidgetItem(truck['truck_type']))
            self.trucks_table.setItem(row, 2, QTableWidgetItem(truck['model']))
            self.trucks_table.setItem(row, 3, QTableWidgetItem(truck['ownership']))
    
    def update_companies_table(self):
        self.companies_table.setRowCount(len(self.companies_data))
        
        for row, company in enumerate(self.companies_data):
            self.companies_table.setItem(row, 0, QTableWidgetItem(company['company_name']))
            self.companies_table.setItem(row, 1, QTableWidgetItem(company['country']))
            self.companies_table.setItem(row, 2, QTableWidgetItem(company['client_name']))
            self.companies_table.setItem(row, 3, QTableWidgetItem(company['deal_type']))
            self.companies_table.setItem(row, 4, QTableWidgetItem(company['address']))
    
    def update_expenses_table(self):
        self.expenses_table.setRowCount(len(self.expenses_data))
        
        for row, expense in enumerate(self.expenses_data):
            self.expenses_table.setItem(row, 0, QTableWidgetItem(expense['shipment_number']))
            self.expenses_table.setItem(row, 1, QTableWidgetItem(str(expense['driver_salary'])))
            self.expenses_table.setItem(row, 2, QTableWidgetItem(str(expense['fuel_cost'])))
            self.expenses_table.setItem(row, 3, QTableWidgetItem(str(expense['oil_cost'])))
            self.expenses_table.setItem(row, 4, QTableWidgetItem(str(expense['maintenance_cost'])))
            self.expenses_table.setItem(row, 5, QTableWidgetItem(str(expense['army_card_cost'])))
            self.expenses_table.setItem(row, 6, QTableWidgetItem(str(expense['rental_cost'])))
            self.expenses_table.setItem(row, 7, QTableWidgetItem(str(expense['delay_fine'])))
            self.expenses_table.setItem(row, 8, QTableWidgetItem(str(expense['transport_cost'])))
    
    
    def update_reports_table(self):
        """Update the reports table with current data"""
        # Check if table exists before trying to update it
        if not hasattr(self, 'reports_table') or self.reports_table is None:
            return
        
        # Clear existing data
        self.reports_table.setRowCount(0)
        
        # Check if we have trip data
        if not hasattr(self, 'trips_data') or not self.trips_data:
            return
        
        # Check if we have expenses data
        if not hasattr(self, 'expenses_data'):
            self.expenses_data = []
        
        self.reports_table.setRowCount(len(self.trips_data))
        
        total_all_expenses = 0.0
        total_all_revenue = 0.0
        total_net_profit = 0.0
        
        for row, trip in enumerate(self.trips_data):
            try:
                shipment_number = trip['shipment_number']
                expenses = [e for e in self.expenses_data if e['shipment_number'] == shipment_number]
                
                total_expenses = 0.0
                revenue = 0.0
                
                if expenses:
                    for expense in expenses:
                        # Safely get expense values with default 0.0
                        expense_items = [
                            expense.get('fuel_cost', 0.0),
                            expense.get('oil_cost', 0.0),
                            expense.get('maintenance_cost', 0.0),
                            expense.get('army_card_cost', 0.0),
                            expense.get('rental_cost', 0.0),
                            expense.get('driver_salary', 0.0),
                            expense.get('delay_fine', 0.0)
                        ]
                        
                        # Convert to float and sum
                        total_expenses += sum(float(item) if item is not None else 0.0 for item in expense_items)
                        revenue += float(expense.get('transport_cost', 0.0))
                
                net_profit = revenue - total_expenses
                
                # Set table items
                self.reports_table.setItem(row, 0, QTableWidgetItem(str(shipment_number)))
                self.reports_table.setItem(row, 1, QTableWidgetItem(f"{total_expenses:.2f}"))
                self.reports_table.setItem(row, 2, QTableWidgetItem(f"{revenue:.2f}"))
                self.reports_table.setItem(row, 3, QTableWidgetItem(f"{net_profit:.2f}"))
                
                total_all_expenses += total_expenses
                total_all_revenue += revenue
                total_net_profit += net_profit
                
            except Exception as e:
                print(f"Error processing trip {row}: {e}")
                # Add empty row in case of error
                for col in range(4):
                    self.reports_table.setItem(row, col, QTableWidgetItem(""))
        
        # Add total row at the bottom
        try:
            total_row = self.reports_table.rowCount()
            self.reports_table.setRowCount(total_row + 1)
            
            # Style the total row differently
            total_item = QTableWidgetItem("الإجمالي")
            total_item.setBackground(QColor(240, 240, 240))  # Light gray background
            self.reports_table.setItem(total_row, 0, total_item)
            
            expenses_item = QTableWidgetItem(f"{total_all_expenses:.2f}")
            expenses_item.setBackground(QColor(240, 240, 240))
            self.reports_table.setItem(total_row, 1, expenses_item)
            
            revenue_item = QTableWidgetItem(f"{total_all_revenue:.2f}")
            revenue_item.setBackground(QColor(240, 240, 240))
            self.reports_table.setItem(total_row, 2, revenue_item)
            
            profit_item = QTableWidgetItem(f"{total_net_profit:.2f}")
            profit_item.setBackground(QColor(240, 240, 240))
            self.reports_table.setItem(total_row, 3, profit_item)
            
        except Exception as e:
            print(f"Error adding total row: {e}")

    def create_reports_content(self):
        """Create the reports content widget"""
        widget = QWidget()
        layout = QVBoxLayout(widget)
        layout.setContentsMargins(20, 20, 20, 20)

        # Header section with user info
        header_layout = QHBoxLayout()
        user_label = QLabel(f"{self.current_user['username']} ({self.current_user['type']})")
        user_label.setAlignment(Qt.AlignRight)
        user_label.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: bold;
                color: #2d3748;
            }
        """)
        header_layout.addStretch()
        header_layout.addWidget(user_label)
        layout.addLayout(header_layout)
        
        # Title and toolbar
        toolbar_layout = QHBoxLayout()
        reports_title = QLabel("تقارير الأرباح والخسائر")
        reports_title.setStyleSheet("""
            QLabel {
                font-size: 24px;
                font-weight: bold;
                color: #2d3748;
                margin: 20px 0;
            }
        """)
        toolbar_layout.addWidget(reports_title)
        toolbar_layout.addStretch()
        
        # Refresh button
        refresh_btn = QPushButton("تحديث التقارير")
        refresh_btn.clicked.connect(self.update_reports_table)
        refresh_btn.setStyleSheet("""
            QPushButton {
                background-color: #4299e1;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #3182ce;
            }
        """)
        toolbar_layout.addWidget(refresh_btn)
        
        # Print button
        print_reports_btn = QPushButton("طباعة")
        print_reports_btn.clicked.connect(self.print_reports_table)
        print_reports_btn.setStyleSheet("""
            QPushButton {
                background-color: #38a169;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #2f855a;
            }
        """)
        toolbar_layout.addWidget(print_reports_btn)
        
        # Export to Excel button
        export_reports_btn = QPushButton("تصدير إلى Excel")
        export_reports_btn.clicked.connect(self.export_reports_to_excel)
        export_reports_btn.setStyleSheet("""
            QPushButton {
                background-color: #ecc94b;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #d69e2e;
            }
        """)
        toolbar_layout.addWidget(export_reports_btn)
        
        layout.addLayout(toolbar_layout)

        # Create or reuse the reports table
        self.create_reports_table()
        layout.addWidget(self.reports_table)
        
        return widget

    def print_reports_table(self):
        """Print the reports table"""
        try:
            printer = QPrinter()
            dialog = QPrintDialog(printer, self)
            
            if dialog.exec_() == QPrintDialog.Accepted:
                painter = QPainter()
                painter.begin(printer)
                
                # Print title
                painter.setFont(QFont("Arial", 16, QFont.Bold))
                painter.drawText(100, 100, "تقارير الأرباح والخسائر")
                
                # Print table content
                if hasattr(self, 'reports_table') and self.reports_table is not None:
                    self.reports_table.render(painter, QRect(50, 200, printer.width()-100, printer.height()-300))
                
                painter.end()
                
        except Exception as e:
            from PyQt5.QtWidgets import QMessageBox
            QMessageBox.warning(self, "خطأ في الطباعة", f"حدث خطأ أثناء الطباعة: {str(e)}")

   ###
    def export_reports_to_excel(self):
        """Export reports table to Excel file"""
        try:   
            if not hasattr(self, 'reports_table') or self.reports_table is None:
                QMessageBox.warning(self, "تحذير", "لا يوجد جدول تقارير للتصدير")
                return
            
            # Get save location
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "حفظ تقارير الأرباح والخسائر", 
                "تقارير_الأرباح_والخسائر.xlsx", 
                "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "تقارير الأرباح والخسائر"
            
            # Add title in a single cell (A1) without merging
            title_cell = ws['A1']
            title_cell.value = "تقارير الأرباح والخسائر"
            title_cell.font = Font(size=16, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            # Add headers
            headers = ["رقم الشحنة", "إجمالي المصاريف", "الإيرادات", "صافي الربح"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
            
            # Add data
            # Replace the data addition loop
            total_all_expenses = 0.0
            total_all_revenue = 0.0
            total_net_profit = 0.0

            for row in range(self.reports_table.rowCount()):
                if self.reports_table.item(row, 0) and self.reports_table.item(row, 0).text() == "الإجمالي":
                    continue  # Skip the total row when reading data
                for col in range(self.reports_table.columnCount()):
                    item = self.reports_table.item(row, col)
                    if item:
                        cell = ws.cell(row=row+4, column=col+1)
                        cell.value = item.text()
                        if col == 1:  # إجمالي المصاريف
                            total_all_expenses += float(item.text())
                        elif col == 2:  # الإيرادات
                            total_all_revenue += float(item.text())
                        elif col == 3:  # صافي الربح
                            total_net_profit += float(item.text())

            # Add total row at the end
            total_row = self.reports_table.rowCount() + 4
            ws.cell(row=total_row, column=1).value = "الإجمالي"
            ws.cell(row=total_row, column=1).font = Font(bold=True)
            ws.cell(row=total_row, column=1).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            ws.cell(row=total_row, column=2).value = f"{total_all_expenses:.2f}"
            ws.cell(row=total_row, column=2).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            ws.cell(row=total_row, column=3).value = f"{total_all_revenue:.2f}"
            ws.cell(row=total_row, column=3).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            ws.cell(row=total_row, column=4).value = f"{total_net_profit:.2f}"
            ws.cell(row=total_row, column=4).fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                        
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            wb.save(file_path)
            QMessageBox.information(self, "نجح التصدير", f"تم تصدير التقارير بنجاح إلى:\n{file_path}")
            
        except ImportError:
            QMessageBox.warning(self, "خطأ", "يجب تثبيت مكتبة openpyxl لتصدير ملفات Excel")
        except Exception as e:
            QMessageBox.critical(self, "خطأ في التصدير", f"حدث خطأ أثناء التصدير: {str(e)}")
   ##

    
    def print_table(self, table, title):
        printer = QPrinter(QPrinter.HighResolution)
        dialog = QPrintDialog(printer, self)
        if dialog.exec_() == QPrintDialog.Accepted:
            painter = QPainter(printer)
            painter.setRenderHint(QPainter.Antialiasing)
            
            painter.setFont(QFont("Arial", 16, QFont.Bold))
            painter.drawText(100, 100, title)
            
            headers = [table.horizontalHeaderItem(i).text() for i in range(table.columnCount())]
            y = 150
            x = 100
            painter.setFont(QFont("Arial", 12, QFont.Bold))
            for header in headers:
                painter.drawText(x, y, header)
                x += 150
            y += 30
            
            painter.setFont(QFont("Arial", 10))
            for row in range(table.rowCount()):
                x = 100
                for col in range(table.columnCount()):
                    item = table.item(row, col)
                    text = item.text() if item else ""
                    painter.drawText(x, y, text)
                    x += 150
                y += 30
            
            painter.end()
            QMessageBox.information(self, 'نجح', 'تمت الطباعة بنجاح!')
    
    def export_table_to_excel_openpyxl(self, table, filename):
        """Export table data to Excel using openpyxl directly with reversed column order for Arabic"""
        try:
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "البيانات"
            
            # Get total number of columns
            total_columns = table.columnCount()
            
            # Add headers in reversed order (last column becomes first, etc.)
            for col in range(total_columns):
                # العمود الأخير في الجدول يصبح العمود الأول في الإكسيل
                reversed_col = total_columns - 1 - col
                header_item = table.horizontalHeaderItem(reversed_col)
                if header_item:
                    cell = ws.cell(row=1, column=col+1)
                    cell.value = header_item.text()
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
            
            # Add data in reversed order
            for row in range(table.rowCount()):
                for col in range(total_columns):
                    # العمود الأخير في الجدول يصبح العمود الأول في الإكسيل
                    reversed_col = total_columns - 1 - col
                    item = table.item(row, reversed_col)
                    if item:
                        cell = ws.cell(row=row+2, column=col+1)  # +2 because row 1 is headers
                        cell.value = item.text()
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            wb.save(filename)
            QMessageBox.information(self, 'نجح', f'تم تصدير البيانات إلى {filename} بنجاح!')
            
        except ImportError:
            QMessageBox.warning(self, "خطأ", "يجب تثبيت مكتبة openpyxl لتصدير ملفات Excel")
        except Exception as e:
            QMessageBox.critical(self, "خطأ في التصدير", f"حدث خطأ أثناء التصدير: {str(e)}")
        
    def print_trips_table(self):
        self.print_table(self.trips_table, "الشحنات الحالية")
    
    def print_drivers_table(self):
        self.print_table(self.drivers_table, "السائقين الحاليين")
    
    def print_trucks_table(self):
        self.print_table(self.trucks_table, "الشاحنات الحالية")
    
    def print_companies_table(self):
        self.print_table(self.companies_table, "الشركات الحالية")
    
    def print_expenses_table(self):
        self.print_table(self.expenses_table, "المصاريف الحالية")
    
    def print_reports_table(self):
        self.print_table(self.reports_table, "التقارير المالية")
    
    def export_trips_to_excel(self):
        """Export trips table to Excel"""
        try:            
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "حفظ بيانات الشحنات", 
                "بيانات_الشحنات.xlsx", 
                "Excel Files (*.xlsx)"
            )
            
            if file_path:
                # Use the openpyxl method for better reliability
                self.export_table_to_excel_openpyxl(self.trips_table, file_path)
                
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ: {str(e)}")
    
    def export_drivers_to_excel(self):
        """Export drivers table to Excel with reversed column order"""
        try:            
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "حفظ بيانات السائقين", 
                "بيانات_السائقين.xlsx", 
                "Excel Files (*.xlsx)"
            )
            
            if file_path:
                self.export_table_to_excel_openpyxl(self.drivers_table, file_path)
                
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ: {str(e)}")

    def export_trucks_to_excel(self):
        """Export trucks table to Excel with reversed column order"""
        try:
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "حفظ بيانات الشاحنات", 
                "بيانات_الشاحنات.xlsx", 
                "Excel Files (*.xlsx)"
            )
            
            if file_path:
                self.export_table_to_excel_openpyxl(self.trucks_table, file_path)
                
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ: {str(e)}")

    def export_companies_to_excel(self):
        """Export companies table to Excel with reversed column order"""
        try:
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "حفظ بيانات الشركات", 
                "بيانات_الشركات.xlsx", 
                "Excel Files (*.xlsx)"
            )
            
            if file_path:
                self.export_table_to_excel_openpyxl(self.companies_table, file_path)
                
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ: {str(e)}")

    def export_expenses_to_excel(self):
        """Export expenses table to Excel with reversed column order"""
        try:
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, 
                "حفظ بيانات المصروفات", 
                "بيانات_المصروفات.xlsx", 
                "Excel Files (*.xlsx)"
            )
            
            if file_path:
                self.export_table_to_excel_openpyxl(self.expenses_table, file_path)
                
        except Exception as e:
            QMessageBox.critical(self, "خطأ", f"حدث خطأ: {str(e)}")
    
   
    def clear_main_content(self):
        for i in reversed(range(self.main_content_layout.count())):
            widget = self.main_content_layout.itemAt(i).widget()
            if widget:
                widget.setVisible(False)  # Hide instead of removing
    
    def show_trips(self):
        self.clear_main_content()
        self.main_content_layout.addWidget(self.create_trips_content())
        self.current_view = 'trips'
    
    def show_drivers(self):
        self.clear_main_content()
        self.main_content_layout.addWidget(self.create_drivers_content())
        self.current_view = 'drivers'
    
    def manage_countries(self):
        dialog = CountryManagerDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            self.update_all_tables()
    
    def show_trucks(self):
        self.clear_main_content()
        self.main_content_layout.addWidget(self.create_trucks_content())
        self.current_view = 'trucks'
    
    def show_companies(self):
        self.clear_main_content()
        self.main_content_layout.addWidget(self.create_companies_content())
        self.current_view = 'companies'
    
    def show_expenses(self):
        self.clear_main_content()
        self.main_content_layout.addWidget(self.create_expenses_content())
        self.current_view = 'expenses'
    
    def show_reports(self):
        if not self.current_user['permissions'].get('reports', False):
            QMessageBox.warning(self, 'خطأ', 'غير مسموح لك بالوصول إلى التقارير!')
            return
            
        self.clear_main_content()
        self.main_content_layout.addWidget(self.create_reports_content())
        self.current_view = 'reports'
    
    def show_settings(self):
        try:
            self.clear_main_content()
            if self.current_user['type'] == 'Admin':
                # تحميل المستخدمين من قاعدة البيانات أولاً
                users = self.load_users()  # استخدام الدالة الموجودة بالفعل
                dialog = UserManagementDialog(users, self)
                dialog.db_file = self.db_file  # تمرير مسار قاعدة البيانات
                if dialog.exec_() == QDialog.Accepted:
                    self.update_all_tables()
            else:
                QMessageBox.warning(self, 'خطأ', 'غير مسموح لك بالوصول إلى إعدادات المستخدمين!')
                
        except Exception as e:
            print(f"خطأ في فتح الإعدادات: {str(e)}")
            QMessageBox.critical(self, 'خطأ', f'حدث خطأ أثناء فتح الإعدادات: {str(e)}')
    
    def update_all_tables(self):
        if hasattr(self, 'trips_table'):
            self.update_trips_table()
        if hasattr(self, 'drivers_table'):
            self.update_drivers_table()
        if hasattr(self, 'trucks_table'):
            self.update_trucks_table()
        if hasattr(self, 'companies_table'):
            self.update_companies_table()
        if hasattr(self, 'expenses_table'):
            self.update_expenses_table()
        if hasattr(self, 'reports_table'):
            self.update_reports_table()


    
    def save_trips_data(self):
        """حفظ بيانات الشحنات في قاعدة البيانات"""
        conn = pymysql.connect(**self.db_file)
        cursor = conn.cursor()
        
        # حذف جميع البيانات القديمة
        cursor.execute("DELETE FROM trips")
        
        # إدراج البيانات الجديدة
        for trip in self.trips_data:
            cursor.execute(
                """INSERT INTO trips 
                (shipment_number, driver_name, shipment_type, shipment_date, 
                start, final_destination, intermediate_country, 
                shipment_status, allowance_period, notes)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (trip['shipment_number'], trip['driver_name'], trip['shipment_type'],
                trip['shipment_date'], trip['start'], trip['final_destination'],
                trip['intermediate_country'], trip['shipment_status'],
                int(trip['allowance_period']), trip['notes'])
            )
        
        conn.commit()
        conn.close()

    def add_new_trip(self):
        dialog = AddTripDialog(self.countries, self.drivers_data, self)
        if dialog.exec_() == QDialog.Accepted:
            trip_data = dialog.get_trip_data()
            if trip_data:
                self.trips_data.append(trip_data)
                self.save_trips_data()  # حفظ في قاعدة البيانات
                self.update_trips_table()
                QMessageBox.information(self, 'نجح', 'تم إضافة الشحنة بنجاح!')

    def edit_trip(self):
        selected_row = self.trips_table.currentRow()
        if selected_row >= 0:
            trip_data = self.trips_data[selected_row]
            dialog = AddTripDialog(self.countries, self.drivers_data, self)
            dialog = AddTripDialog(self.countries, self.drivers_data, self)
            dialog.shipment_number.setText(trip_data['shipment_number'])
            dialog.driver_name.setCurrentText(trip_data['driver_name'])
            dialog.shipment_type.setCurrentText(trip_data['shipment_type'])
            dialog.shipment_date.setDate(QDate.fromString(trip_data['shipment_date'], 'yyyy-MM-dd'))
            dialog.start.setCurrentText(trip_data.get('start', ''))  # تعيين قيمة "مسافرة من"
            dialog.final_destination.setCurrentText(trip_data['final_destination'])
            dialog.intermediate_country.setCurrentText(trip_data['intermediate_country'] or "لا يوجد")
            dialog.shipment_status.setCurrentText(trip_data['shipment_status'])
            dialog.allowance_period.setText(trip_data['allowance_period'])
            dialog.notes.setText(trip_data['notes'])
            if dialog.exec_() == QDialog.Accepted:
                new_trip_data = dialog.get_trip_data()
                if new_trip_data:
                    self.trips_data[selected_row] = new_trip_data
                    self.save_trips_data()  # حفظ في قاعدة البيانات
                    self.update_trips_table()
                    QMessageBox.information(self, 'نجح', 'تم تعديل الشحنة بنجاح!')

    def delete_trip(self):
        selected_row = self.trips_table.currentRow()
        if selected_row >= 0:
            shipment_number = self.trips_data[selected_row]['shipment_number']
            reply = QMessageBox.question(self, 'تأكيد الحذف', 'هل تريد حذف هذه الشحنة وصورها؟')
            if reply == QMessageBox.Yes:
                try:
                    # حذف الصور المرتبطة
                    conn = pymysql.connect(**self.db_file)
                    cursor = conn.cursor()
                    cursor.execute("SELECT image_path FROM shipment_images WHERE shipment_number = %s", (shipment_number,))
                    for (image_path,) in cursor.fetchall():
                        if os.path.exists(image_path):
                            os.remove(image_path)
                    cursor.execute("DELETE FROM shipment_images WHERE shipment_number = %s", (shipment_number,))
                    conn.commit()
                    conn.close()

                    # حذف الشحنة
                    del self.trips_data[selected_row]
                    self.save_trips_data()
                    self.update_trips_table()
                    QMessageBox.information(self, 'نجح', 'تم حذف الشحنة وصورها بنجاح!')
                    
                except Exception as e:
                    print(f"Error deleting trip and images: {str(e)}")
                    QMessageBox.critical(self, 'خطأ', f'حدث خطأ أثناء حذف الشحنة: {str(e)}')
        

    def save_drivers_data(self):
        """Save drivers data to the database"""
        conn = pymysql.connect(**self.db_file)
        cursor = conn.cursor()
        
        # Delete all existing drivers
        cursor.execute("DELETE FROM drivers")
        
        # Insert current drivers
        for driver in self.drivers_data:
            cursor.execute(
                "INSERT INTO drivers (name, address, age, national_id) VALUES (%s, %s, %s, %s)",
                (driver['name'], driver['address'], int(driver['age']) if driver['age'] else None, driver['national_id'])
            )
        
        conn.commit()
        conn.close()
    
    def add_new_driver(self):
        dialog = AddDriverDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            driver_data = dialog.get_driver_data()
            self.drivers_data.append(driver_data)
            self.save_drivers_data()  # حفظ في قاعدة البيانات
            self.update_drivers_table()
            QMessageBox.information(self, 'نجح', 'تم إضافة السائق بنجاح!')

    def edit_driver(self):
        selected_row = self.drivers_table.currentRow()
        if selected_row >= 0:
            driver_data = self.drivers_data[selected_row]
            dialog = AddDriverDialog(self)
            dialog.name.setText(driver_data['name'])
            dialog.address.setText(driver_data['address'])
            dialog.age.setText(driver_data['age'])
            dialog.national_id.setText(driver_data['national_id'])
            if dialog.exec_() == QDialog.Accepted:
                new_driver_data = dialog.get_driver_data()
                self.drivers_data[selected_row] = new_driver_data
                self.save_drivers_data()  # حفظ في قاعدة البيانات
                self.update_drivers_table()
                QMessageBox.information(self, 'نجح', 'تم تعديل السائق بنجاح!')

    def delete_driver(self):
        selected_row = self.drivers_table.currentRow()
        if selected_row >= 0:
            reply = QMessageBox.question(self, 'تأكيد الحذف', 'هل تريد حذف هذا السائق؟')
            if reply == QMessageBox.Yes:
                del self.drivers_data[selected_row]
                self.save_drivers_data()  # حفظ في قاعدة البيانات
                self.update_drivers_table()
                QMessageBox.information(self, 'نجح', 'تم حذف السائق بنجاح!')
    
    def save_trucks_data(self):
        """حفظ بيانات الشاحنات في قاعدة البيانات"""
        conn = pymysql.connect(**self.db_file)
        cursor = conn.cursor()
        
        cursor.execute("DELETE FROM trucks")
        
        for truck in self.trucks_data:
            cursor.execute(
                "INSERT INTO trucks (truck_number, truck_type, model, ownership) VALUES (%s, %s, %s, %s)",
                (truck['truck_number'], truck['truck_type'], 
                truck['model'], truck['ownership'])
            )
        
        conn.commit()
        conn.close()

    def add_new_truck(self):
        dialog = AddTruckDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            truck_data = dialog.get_truck_data()
            self.trucks_data.append(truck_data)
            self.save_trucks_data()  # حفظ في قاعدة البيانات
            self.update_trucks_table()
            QMessageBox.information(self, 'نجح', 'تم إضافة الشاحنة بنجاح!')

    def edit_truck(self):
        selected_row = self.trucks_table.currentRow()
        if selected_row >= 0:
            truck_data = self.trucks_data[selected_row]
            dialog = AddTruckDialog(self)
            # تعيين قيم الحوار...
            dialog = AddTruckDialog(self)
            dialog.truck_number.setText(truck_data['truck_number'])
            dialog.truck_type.setText(truck_data['truck_type'])
            dialog.model.setText(truck_data['model'])
            dialog.ownership.setCurrentText(truck_data['ownership'])
            if dialog.exec_() == QDialog.Accepted:
                new_truck_data = dialog.get_truck_data()
                self.trucks_data[selected_row] = new_truck_data
                self.save_trucks_data()  # حفظ في قاعدة البيانات
                self.update_trucks_table()
                QMessageBox.information(self, 'نجح', 'تم تعديل الشاحنة بنجاح!')

    def delete_truck(self):
        selected_row = self.trucks_table.currentRow()
        if selected_row >= 0:
            reply = QMessageBox.question(self, 'تأكيد الحذف', 'هل تريد حذف هذه الشاحنة؟')
            if reply == QMessageBox.Yes:
                del self.trucks_data[selected_row]
                self.save_trucks_data()  # حفظ في قاعدة البيانات
                self.update_trucks_table()
                QMessageBox.information(self, 'نجح', 'تم حذف الشاحنة بنجاح!')
        
    def save_companies_data(self):
        """حفظ بيانات الشركات في قاعدة البيانات"""
        conn = pymysql.connect(**self.db_file)
        cursor = conn.cursor()
        
        cursor.execute("DELETE FROM companies")
        
        for company in self.companies_data:
            cursor.execute(
                """INSERT INTO companies 
                (company_name, country, client_name, deal_type, address)
                VALUES (%s, %s, %s, %s, %s)""",
                (company['company_name'], company['country'], 
                company['client_name'], company['deal_type'], 
                company['address'])
            )
        
        conn.commit()
        conn.close()

    def add_new_company(self):
        dialog = AddCompanyDialog(self.countries, self)
        if dialog.exec_() == QDialog.Accepted:
            company_data = dialog.get_company_data()
            self.companies_data.append(company_data)
            self.save_companies_data()  # حفظ في قاعدة البيانات
            self.update_companies_table()
            QMessageBox.information(self, 'نجح', 'تم إضافة الشركة بنجاح!')

    def edit_company(self):
        selected_row = self.companies_table.currentRow()
        if selected_row >= 0:
            company_data = self.companies_data[selected_row]
            # تعيين قيم الحوار...

            dialog = AddCompanyDialog(self.countries, self)
            dialog.company_name.setText(company_data['company_name'])
            dialog.country.setCurrentText(company_data['country'])
            dialog.client_name.setText(company_data['client_name'])
            dialog.deal_type.setCurrentText(company_data['deal_type'])
            dialog.address.setText(company_data['address'])
            if dialog.exec_() == QDialog.Accepted:
                new_company_data = dialog.get_company_data()
                self.companies_data[selected_row] = new_company_data
                self.save_companies_data()  # حفظ في قاعدة البيانات
                self.update_companies_table()
                QMessageBox.information(self, 'نجح', 'تم تعديل الشركة بنجاح!')

    def delete_company(self):
        selected_row = self.companies_table.currentRow()
        if selected_row >= 0:
            reply = QMessageBox.question(self, 'تأكيد الحذف', 'هل تريد حذف هذه الشركة؟')
            if reply == QMessageBox.Yes:
                del self.companies_data[selected_row]
                self.save_companies_data()  # حفظ في قاعدة البيانات
                self.update_companies_table()
                QMessageBox.information(self, 'نجح', 'تم حذف الشركة بنجاح!')
    
    def save_expenses_data(self):
        """حفظ بيانات المصاريف في قاعدة البيانات"""
        conn = pymysql.connect(**self.db_file)
        cursor = conn.cursor()
        
        cursor.execute("DELETE FROM expenses")
        
        for expense in self.expenses_data:
            cursor.execute(
                """INSERT INTO expenses 
                (shipment_number, fuel_cost, oil_cost, maintenance_cost, 
                army_card_cost, rental_cost, driver_salary, delay_fine, transport_cost)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                (expense['shipment_number'], 
                float(expense['fuel_cost']), float(expense['oil_cost']),
                float(expense['maintenance_cost']), float(expense['army_card_cost']),
                float(expense['rental_cost']), float(expense['driver_salary']),
                float(expense['delay_fine']), float(expense['transport_cost']))
            )
        
        conn.commit()
        conn.close()

    def add_new_expense(self):
        if not self.trips_data:
            QMessageBox.warning(self, 'خطأ', 'يرجى إضافة شحنة أولاً!')
            return
            
        dialog = AddExpenseDialog(self.trips_data, self)
        if dialog.exec_() == QDialog.Accepted:
            expense_data = dialog.get_expense_data()
            self.expenses_data.append(expense_data)
            self.save_expenses_data()  # حفظ في قاعدة البيانات
            self.update_expenses_table()
            if hasattr(self, 'reports_table'):
                self.update_reports_table()
            QMessageBox.information(self, 'نجح', 'تم إضافة المصاريف بنجاح!')

    def edit_expense(self):
        selected_row = self.expenses_table.currentRow()
        if selected_row >= 0:
            expense_data = self.expenses_data[selected_row]
            dialog = AddExpenseDialog(self.trips_data, self)
            # تعيين قيم الحوار...
            dialog.shipment_number.setCurrentText(expense_data['shipment_number'])
            dialog.fuel_cost.setText(str(expense_data['fuel_cost']))
            dialog.oil_cost.setText(str(expense_data['oil_cost']))
            dialog.maintenance_cost.setText(str(expense_data['maintenance_cost']))
            dialog.army_card_cost.setText(str(expense_data['army_card_cost']))
            dialog.rental_cost.setText(str(expense_data['rental_cost']))
            dialog.driver_salary.setText(str(expense_data['driver_salary']))
            dialog.delay_fine.setText(str(expense_data['delay_fine']))
            dialog.transport_cost.setText(str(expense_data['transport_cost']))
            if dialog.exec_() == QDialog.Accepted:
                new_expense_data = dialog.get_expense_data()
                self.expenses_data[selected_row] = new_expense_data
                self.save_expenses_data()  # حفظ في قاعدة البيانات
                self.update_expenses_table()
                if hasattr(self, 'reports_table'):
                    self.update_reports_table()
                QMessageBox.information(self, 'نجح', 'تم تعديل المصاريف بنجاح!')

    def delete_expense(self):
        selected_row = self.expenses_table.currentRow()
        if selected_row >= 0:
            reply = QMessageBox.question(self, 'تأكيد الحذف', 'هل تريد حذف هذه المصاريف؟')
            if reply == QMessageBox.Yes:
                del self.expenses_data[selected_row]
                self.save_expenses_data()  # حفظ في قاعدة البيانات
                self.update_expenses_table()
                if hasattr(self, 'reports_table'):
                    self.update_reports_table()
                QMessageBox.information(self, 'نجح', 'تم حذف المصاريف بنجاح!')
if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = ShippingManagementApp()
    window.show()
    sys.exit(app.exec_())
